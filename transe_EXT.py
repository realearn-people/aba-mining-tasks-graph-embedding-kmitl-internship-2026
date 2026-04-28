import pandas as pd
import numpy as np
import torch
import time
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pykeen.pipeline import pipeline
from pykeen.triples import TriplesFactory
from pykeen.sampling import BasicNegativeSampler
from pykeen.evaluation import RankBasedEvaluator
from sklearn.decomposition import PCA

# ── File paths ────────────────────────────────────────────────────────────────
CSV_PATH     = "hotel_contrary_dataset_all.csv"   # input dataset
OUTPUT_DIR   = Path("transe_output")               # folder for saved outputs
RESULTS_FILE = "experiment_results.xlsx"           # auto-recorded experiment log
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Hyperparameters ───────────────────────────────────────────────────────────
DATASET       = "ALL"      # dataset label for experiment record
EMBEDDING_DIM = 100        # size of each entity/relation vector
NUM_EPOCHS    = 1000       # max epochs — early stopper will stop earlier
BATCH_SIZE    = 256        # number of triples processed per update step
LEARNING_RATE = 0.0001     # how much embeddings shift per step (Adam optimizer)
MARGIN        = 1.0        # minimum score gap between positive and negative triples
TRAIN_RATIO   = 0.8        # 80% of triples used for training
VAL_RATIO     = 0.1        # 10% used for validation (early stopping monitor), remaining 10% used for final test evaluation
RANDOM_SEED   = 42         # fixes randomness for reproducibility
NUM_NEG       = 2          # negatives per positive triple
MODEL         = "TransE"   # rotation-based KGE model — handles symmetric relations
LOSS          = "NSSALoss" # Non-Saturating Self-Adversarial Loss weights harder negatives more — better for relation corruption
ADV_TEMP      = 0.5        # adversarial temperature for NSSALoss
CORRUPT       = "relation" # corruption scheme — only replace the relation

# ── Extended negative sampler ─────────────────────────────────────────────────
class ExtendedBasicNegativeSampler(BasicNegativeSampler):
    """
    Custom negative sampler that extends BasicNegativeSampler.
    Overrides corrupt_batch() to capture and save every negative triple
    generated during training to a CSV file.

    Each batch writes immediately to disk — no accumulation in RAM —
    to avoid memory overflow on large datasets.

    Based on PyKEEN GitHub issue #1124:
    https://github.com/pykeen/pykeen/issues/1124#issuecomment-1262011487
    """

    def __init__(self, *args, save_path=str(OUTPUT_DIR / "negative_samples.csv"), **kwargs):
        # Force relation-only corruption inside the class
        # Cannot rely on PyKEEN passing corruption_scheme correctly to custom classes
        kwargs['corruption_scheme'] = ['relation']
        super().__init__(*args, **kwargs)

        # Use object.__setattr__ to bypass PyTorch nn.Module's attribute interception
        # nn.Module overrides __setattr__ for parameter management
        # which would reject plain Python attributes like these
        object.__setattr__(self, 'save_path', save_path)
        object.__setattr__(self, 'header_written', False)  # tracks if CSV header exists
        object.__setattr__(self, 'batch_count', 0)         # counts batches processed

    def corrupt_batch(self, positive_batch: torch.LongTensor) -> torch.LongTensor:
        """
        Called automatically by PyKEEN for every batch during training.
        Generates negatives using parent class logic, then saves them to CSV.
        Returns negatives unchanged so training proceeds normally.
        """
        # Generate negatives using BasicNegativeSampler's random relation corruption
        negatives = super().corrupt_batch(positive_batch=positive_batch)

        # Increment batch counter using object.__setattr__ (same reason as __init__)
        object.__setattr__(self, 'batch_count', self.batch_count + 1)

        # Debug messages — confirm corrupt_batch is being called
        if self.batch_count == 1:
            print(f"[DEBUG] corrupt_batch called for the first time!")
            print(f"[DEBUG] Negative shape: {negatives.shape}")
        if self.batch_count % 100 == 0:
            print(f"[DEBUG] Batches captured so far: {self.batch_count}")

        try:
            # Convert integer IDs back to readable phrase labels
            neg_flat = negatives.detach().cpu().reshape(-1, 3)
            rows = []
            for t in neg_flat:
                try:
                    rows.append({
                        "head":     id_to_entity.get(t[0].item(), "unknown"),
                        "relation": id_to_relation.get(t[1].item(), "unknown"),
                        "tail":     id_to_entity.get(t[2].item(), "unknown"),
                        "label":    0,   # 0 = negative (fake) triple
                    })
                except Exception as row_err:
                    print(f"[WARN] Skipping one row in batch {self.batch_count}: {row_err}")
                    continue
            if rows:
                df_batch = pd.DataFrame(rows)
                df_batch.to_csv(
                    self.save_path,
                    mode="w" if not self.header_written else "a",
                    header=not self.header_written,
                    index=False,      
                )
                object.__setattr__(self, 'header_written', True)
                del df_batch
            del neg_flat, rows
        except Exception as batch_err:
            print(f"[WARN] Skipping batch {self.batch_count} due to error: {batch_err}")
        return negatives

# ══════════════════════════════════════════════════════════════════════════════
# Rank-capturing evaluator
# ══════════════════════════════════════════════════════════════════════════════
class RankCapturingEvaluator(RankBasedEvaluator):
    """
    Extends RankBasedEvaluator to capture per-triple ranks during
    process_scores_(), which is called once per batch by PyKEEN.
    Captures both head and tail ranks separately.
    """
 
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._captured_tail_ranks = []
        self._captured_head_ranks = []
 
    def process_scores_(self, hrt_batch, target, scores, true_scores, dense_positive_mask=None):
        super().process_scores_(
            hrt_batch=hrt_batch,
            target=target,
            scores=scores,
            true_scores=true_scores,
            dense_positive_mask=dense_positive_mask,
        )
        try:
            with torch.no_grad():
                # Realistic rank: strictly higher scores + 1
                ranks = (scores > true_scores).sum(dim=1).cpu().numpy() + 1
                ranks = ranks.tolist()
 
            if target == 'tail':
                self._captured_tail_ranks.extend(ranks)
            elif target == 'head':
                self._captured_head_ranks.extend(ranks)
 
        except Exception as e:
            print(f"[WARN] Could not capture ranks for batch (target={target}): {e}")
 

# ── Experiment recorder ───────────────────────────────────────────────────────
def record_results(filepath, settings, metrics, elapsed):
    """
    Appends one row of settings and metrics to the experiment Excel file.
    Creates the file with a formatted header if it doesn't exist yet.
    Run number is auto-incremented based on existing rows.
    """
    headers = [
        "No.", "Timestamp", "Dataset", "Model", "Loss Function",
        "Neg Sampler", "Corruption", "Embedding Dim", "Num Epochs",
        "Best Epoch", "Batch Size", "Learning Rate", "Margin",
        "Adv Temp", "Num Neg", "Training Time",
        "MRR", "Hits@1", "Hits@3", "Hits@10", "Mean Rank",
    ]

    header_fill  = PatternFill("solid", fgColor="0D1B4B")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Load existing file or create new one
    if Path(filepath).exists():
        wb       = load_workbook(filepath)
        ws       = wb.active
        next_row = ws.max_row + 1
        run_num  = next_row - 1   # row 1 is the header
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Experiment Results"
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"    # keep header visible when scrolling
        next_row = 2
        run_num  = 1

    # Build the row with all settings and metrics
    row_data = [
        run_num,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        settings["dataset"],
        settings["model"],
        settings["loss_func"],
        settings["neg_sampler"],
        settings["corrupt"],
        settings["embedding_dim"],
        settings["num_epochs"],
        settings.get("best_epoch", "—"),
        settings["batch_size"],
        settings["learning_rate"],
        settings["margin"],
        settings["adv_temp"],
        settings["num_neg"],
        f"{int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s",
        round(metrics["mrr"],        4),
        round(metrics["hits_at_1"],  4),
        round(metrics["hits_at_3"],  4),
        round(metrics["hits_at_10"], 4),
        round(metrics["mean_rank"],  4),
    ]

    for col, val in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col, value=val)

    wb.save(filepath)
    print(f"Recorded: {filepath}  Run #{run_num}")


# ══════════════════════════════════════════════════════════════════════════════
# Load data
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Loading triples")
print("=" * 60)

df = pd.read_csv(CSV_PATH)
print(f"Total rows loaded: {len(df):,}")
print(f"Relations found: {df['relation'].value_counts().to_dict()}")

# Extract only the 3 columns PyKEEN needs — head, relation, tail
triples_array = df[['head', 'relation', 'tail']].values.astype(str)

# ══════════════════════════════════════════════════════════════════════════════
# Build TriplesFactory and split
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Creating TriplesFactory")
print("=" * 60)

# TriplesFactory converts string phrases to integer IDs for the model
tf = TriplesFactory.from_labeled_triples(triples=triples_array)
print(f"Unique entities  : {tf.num_entities:,}")
print(f"Unique relations : {tf.num_relations}")

# Split into train / validation / test sets
# Test set = remaining 10% after train (80%) and val (10%)
training, validation, testing = tf.split(
    ratios=[TRAIN_RATIO, VAL_RATIO],
    random_state=RANDOM_SEED,
)
print(f"Train : {training.num_triples:,}")
print(f"Val   : {validation.num_triples:,}")
print(f"Test  : {testing.num_triples:,}")

# Build lookup dictionaries — used inside ExtendedBasicNegativeSampler
# to convert integer IDs back to readable phrase labels when saving to CSV
id_to_entity   = {v: k for k, v in training.entity_to_id.items()}
id_to_relation = {v: k for k, v in training.relation_to_id.items()}

# ══════════════════════════════════════════════════════════════════════════════
# Train
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Training " + MODEL)
print("=" * 60)

start_time = time.time()

result = pipeline(
    training=training,
    validation=validation,
    testing=testing,

    # Model — RotatE represents relations as rotations in complex space
    # Naturally handles symmetric relations like CONTRARY_TO
    model=MODEL,
    model_kwargs=dict(embedding_dim=EMBEDDING_DIM),

    # Loss — NSSALoss weights harder negatives more during training
    loss=LOSS,
    loss_kwargs=dict(
        margin=MARGIN,
        adversarial_temperature=ADV_TEMP,
    ),

    # Negative sampler — custom class that saves negatives to CSV during training
    # Pass class (not instance) — PyKEEN instantiates it internally
    # corruption_scheme is forced to ["relation"] inside the class __init__
    negative_sampler=ExtendedBasicNegativeSampler,
    negative_sampler_kwargs=dict(
        num_negs_per_pos=NUM_NEG,
        corruption_scheme=[CORRUPT],   # also passed here as documentation
    ),

    # Optimizer — Adam adapts learning rate per parameter
    # handles sparse gradients well (most entities appear rarely)
    optimizer="Adam",
    optimizer_kwargs=dict(lr=LEARNING_RATE),

    # Training loop — sLCWA samples NUM_NEG negatives per positive per batch
    # more efficient than full LCWA which scores all possible entities
    training_loop="sLCWA",
    training_kwargs=dict(
        num_epochs=NUM_EPOCHS,
        batch_size=BATCH_SIZE,
    ),

    # Early stopping — monitors Hits@10 on validation set every 10 epochs
    # stops if no improvement of at least 0.2% for 10 consecutive checks (100 epochs)
    stopper="early",
    stopper_kwargs=dict(
        metric="hits_at_10",
        patience=10,
        relative_delta=0.002,
        frequency=10,
    ),

    # Evaluation — rank-based: for each test triple, ranks all entities
    # filtered=True removes other known true triples from ranking for fairness
    evaluator="RankBasedEvaluator",
    evaluator_kwargs=dict(filtered=True),

    random_seed=RANDOM_SEED,

    # Automatically use GPU if available, otherwise fall back to CPU
    device="cuda" if torch.cuda.is_available() else "cpu",
)

end_time = time.time()
elapsed  = end_time - start_time
print(f"\nTraining completed in: {int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s")

# ══════════════════════════════════════════════════════════════════════════════
# PyKEEN default evaluation results 
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Evaluation Results (PyKEEN default)")
print("=" * 60)
 
metrics_dict = result.metric_results.to_dict()
r            = metrics_dict["both"]["realistic"]
 
mrr        = r['inverse_harmonic_mean_rank']
hits_at_1  = r['hits_at_1']
hits_at_3  = r['hits_at_3']
hits_at_10 = r['hits_at_10']
mean_rank  = r['arithmetic_mean_rank']
 
print(f"MRR       : {mrr:.4f}")
print(f"Hits@1    : {hits_at_1:.4f}")
print(f"Hits@3    : {hits_at_3:.4f}")
print(f"Hits@10   : {hits_at_10:.4f}")
print(f"Mean Rank : {mean_rank:.4f}")

# ══════════════════════════════════════════════════════════════════════════════
# Record results to Excel
# ══════════════════════════════════════════════════════════════════════════════

# Get the epoch where the best validation result occurred
try:
    best_epoch = result.stopper.best_epoch
except Exception:
    best_epoch = "—"

settings = {
    "dataset":       DATASET,
    "model":         MODEL,
    "loss_func":     LOSS,
    "neg_sampler":   "ExtendedBasicNegativeSampler",
    "corrupt":       CORRUPT,
    "embedding_dim": EMBEDDING_DIM,
    "num_epochs":    NUM_EPOCHS,
    "best_epoch":    best_epoch,
    "batch_size":    BATCH_SIZE,
    "learning_rate": LEARNING_RATE,
    "margin":        MARGIN,
    "adv_temp":      ADV_TEMP,
    "num_neg":       NUM_NEG,
}

metrics_out = {
    "mrr":        mrr,
    "hits_at_1":  hits_at_1,
    "hits_at_3":  hits_at_3,
    "hits_at_10": hits_at_10,
    "mean_rank":  mean_rank,
}

record_results(RESULTS_FILE, settings, metrics_out, elapsed)

print("=" * 60)
print("Saving per-triple prediction ranks")
print("=" * 60)
 
try:
    evaluator = RankCapturingEvaluator(filtered=True)
 
    eval_result = evaluator.evaluate(
        model=result.model,
        mapped_triples=testing.mapped_triples,
        additional_filter_triples=[
            training.mapped_triples,
            validation.mapped_triples,
        ],
        batch_size=BATCH_SIZE,
        device="cuda" if torch.cuda.is_available() else "cpu",
        
    )
    tail_ranks = evaluator._captured_tail_ranks
    head_ranks = evaluator._captured_head_ranks
 
    print(f"Captured tail ranks : {len(tail_ranks)}")
    print(f"Captured head ranks : {len(head_ranks)}")
    print(f"Test triples        : {len(testing.mapped_triples)}")
 
    if len(tail_ranks) == 0 or len(head_ranks) == 0:
        raise ValueError("No ranks captured — process_scores_() may not have been called")
 
    # ── Build per-triple dataframe ────────────────────────────────────────────
    eval_rows    = []
    test_triples = testing.mapped_triples
    n            = min(len(test_triples), len(tail_ranks), len(head_ranks))
 
    for i in range(n):
        try:
            triple = test_triples[i]
            h_name = id_to_entity.get(triple[0].item(), "unknown")
            r_name = id_to_relation.get(triple[1].item(), "unknown")
            t_name = id_to_entity.get(triple[2].item(), "unknown")
 
            t_rank   = int(tail_ranks[i])
            h_rank   = int(head_ranks[i])
            avg_rank = (t_rank + h_rank) / 2
            avg_rr   = (1 / t_rank + 1 / h_rank) / 2
 
            # Hits averaged per-side — matches PyKEEN's 'both' aggregation
            eval_rows.append({
                "head":             h_name,
                "relation":         r_name,
                "tail":             t_name,
                "tail_rank":        t_rank,
                "head_rank":        h_rank,
                "avg_rank":         round(avg_rank, 2),
                "reciprocal_rank":  round(avg_rr, 6),
                "tail_hits_at_1":   int(t_rank <= 1),
                "tail_hits_at_3":   int(t_rank <= 3),
                "tail_hits_at_10":  int(t_rank <= 10),
                "head_hits_at_1":   int(h_rank <= 1),
                "head_hits_at_3":   int(h_rank <= 3),
                "head_hits_at_10":  int(h_rank <= 10),
                "hits_at_1":        (int(t_rank <= 1)  + int(h_rank <= 1))  / 2,
                "hits_at_3":        (int(t_rank <= 3)  + int(h_rank <= 3))  / 2,
                "hits_at_10":       (int(t_rank <= 10) + int(h_rank <= 10)) / 2,
            })
 
        except Exception as e:
            print(f"[WARN] Skipping triple {i}: {e}")
            continue
 
    df_eval   = pd.DataFrame(eval_rows)
    EVAL_FILE = OUTPUT_DIR / "eval_ranked_triples.csv"
 
    # Always overwrite — each run produces fresh per-triple ranks
    df_eval.to_csv(EVAL_FILE, index=False, mode='w')
    print(f"\nSaved (overwritten): {EVAL_FILE}  ({len(df_eval):,} triples)")
 
    # ── Verification — should match PyKEEN reported metrics ───────────────────
    print("\nVerification:")
    print(f"MRR       : {df_eval['reciprocal_rank'].mean():.4f}  (PyKEEN: {mrr:.4f})")
    print(f"Mean Rank : {df_eval['avg_rank'].mean():.4f}  (PyKEEN: {mean_rank:.4f})")
    print(f"Hits@1    : {df_eval['hits_at_1'].mean():.4f}  (PyKEEN: {hits_at_1:.4f})")
    print(f"Hits@3    : {df_eval['hits_at_3'].mean():.4f}  (PyKEEN: {hits_at_3:.4f})")
    print(f"Hits@10   : {df_eval['hits_at_10'].mean():.4f}  (PyKEEN: {hits_at_10:.4f})")
 
except Exception as e:
    print(f"[ERROR] Per-triple evaluation failed: {e}")
    import traceback; traceback.print_exc()
 
 
# ══════════════════════════════════════════════════════════════════════════════
# Export visualization CSV
# RotatE embeddings are complex — concatenate real + imaginary for PCA
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Exporting visualization data")
print("=" * 60)
 
try:
    # TransE embeddings are already real-valued — no complex conversion needed
    entity_emb_real = result.model.entity_representations[0](indices=None)\
                      .detach().cpu().numpy().astype(np.float32)
 
    print(f"Entity embedding shape (real+imag): {entity_emb_real.shape}")
 
    # Project to 2D with PCA
    pca       = PCA(n_components=2, random_state=RANDOM_SEED)
    coords_2d = pca.fit_transform(entity_emb_real)
    print(f"PCA explained variance ratio: {pca.explained_variance_ratio_.round(4)}")
 
    # Build entity dataframe with 2D coordinates
    entity_to_id = training.entity_to_id
    entity_rows  = []
    for entity_name, entity_id in entity_to_id.items():
        entity_rows.append({
            "entity_id":   entity_id,
            "entity_name": entity_name,
            "x":           round(float(coords_2d[entity_id, 0]), 6),
            "y":           round(float(coords_2d[entity_id, 1]), 6),
        })
    df_entities = pd.DataFrame(entity_rows)
 
    # ── Score all original triples ────────────────────────────────────────────
    df_orig = pd.read_csv(CSV_PATH)[['head', 'relation', 'tail']]
 
    scores_list = []
    result.model.eval()
    with torch.no_grad():
        for _, row in df_orig.iterrows():
            try:
                h_id = training.entity_to_id.get(str(row['head']), -1)
                r_id = training.relation_to_id.get(str(row['relation']), -1)
                t_id = training.entity_to_id.get(str(row['tail']), -1)
                if -1 in (h_id, r_id, t_id):
                    scores_list.append(None)
                    continue
                hrt = torch.tensor([[h_id, r_id, t_id]], dtype=torch.long)
                s   = result.model.score_hrt(hrt)
                scores_list.append(round(float(s.item()), 6))
            except Exception:
                scores_list.append(None)
 
    df_orig['score'] = scores_list
 
    # ── Join 2D coordinates for head and tail ─────────────────────────────────
    coord_map         = df_entities.set_index('entity_name')[['x', 'y']]
    df_orig['head_x'] = df_orig['head'].map(coord_map['x'])
    df_orig['head_y'] = df_orig['head'].map(coord_map['y'])
    df_orig['tail_x'] = df_orig['tail'].map(coord_map['x'])
    df_orig['tail_y'] = df_orig['tail'].map(coord_map['y'])
 
    # ── Save both visualization files ─────────────────────────────────────────
    VIZ_FILE    = OUTPUT_DIR / "visualization_data.csv"
    ENTITY_FILE = OUTPUT_DIR / "entity_embeddings.csv"
 
    df_orig.to_csv(VIZ_FILE, index=False)
    df_entities.to_csv(ENTITY_FILE, index=False)
 
    print(f"Saved: {VIZ_FILE}     ({len(df_orig):,} triples)")
    print(f"Saved: {ENTITY_FILE}  ({len(df_entities):,} entities)")
 
except Exception as e:
    print(f"[ERROR] Visualization export failed: {e}")
    import traceback; traceback.print_exc()
 
 
print("=" * 60)
print("All done.")
print("=" * 60)