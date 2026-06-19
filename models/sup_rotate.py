"""
For 3-relation training (CONTRARY_TO, NOT_CONTRARY, SUPPORT)

"""

import pandas as pd
import numpy as np
import torch
import time
import datetime
import pickle
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pykeen.pipeline import pipeline
from pykeen.triples import TriplesFactory
from pykeen.sampling import BasicNegativeSampler
from sklearn.decomposition import PCA
from sklearn.model_selection import train_test_split

# ── File paths ────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
CSV_PATH     = ROOT / "data" / "input_data" / "hotel_contrary_dataset_support.csv"   # input dataset
OUTPUT_DIR   = ROOT / "outputs" / "rotate_sup_output"             # folder for saved outputs
RESULTS_FILE = ROOT / "data" / "experiment_results" / "sup_experiment_results.xlsx"           # auto-recorded experiment log
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Hyperparameters ───────────────────────────────────────────────────────────
DATASET       = "ALL"     # dataset label for experiment record
EMBEDDING_DIM = 100        # size of each entity/relation vector
NUM_EPOCHS    = 1000       # max epochs — early stopper will stop earlier
BATCH_SIZE    = 256        # number of triples processed per update step
LEARNING_RATE = 0.0001     # how much embeddings shift per step (Adam optimizer)
MARGIN        = 1.0        # minimum score gap between positive and negative triples
TRAIN_RATIO   = 0.8        # 80% of triples used for training
VAL_RATIO     = 0.1        # 10% used for validation (early stopping monitor), remaining 10% used for final test evaluation
RANDOM_SEED   = 42         # fixes randomness for reproducibility
NUM_NEG       = 3          # negatives per positive triple
MODEL         = "RotatE"   # rotation-based KGE model — handles symmetric relations
LOSS          = "NSSALoss" # Non-Saturating Self-Adversarial Loss weights harder negatives more — better for relation corruption
ADV_TEMP      = 0.5        # adversarial temperature for NSSALoss
CORRUPT       = "relation" # corruption scheme — only replace the relation

# ── ABA claim roots ───────────────────────────────────────────────────────────
CLAIMS = [
    'good_staff',    'bad_staff',
    'good_price',    'bad_price',
    'good_check-in', 'bad_check-in',
    'good_check-out','bad_check-out',
]

# ── Extended negative sampler ─────────────────────────────────────────────────
class ExtendedBasicNegativeSampler(BasicNegativeSampler):

    def __init__(self, *args, save_path=str(OUTPUT_DIR / "negative_samples.csv"), **kwargs):
        # Force relation-only corruption inside the class
        # Cannot rely on PyKEEN passing corruption_scheme correctly to custom classes
        kwargs['corruption_scheme'] = ['relation']
        super().__init__(*args, **kwargs)

        # Use object.__setattr__ to bypass PyTorch nn.Module's attribute interception
        # nn.Module overrides __setattr__ for parameter management
        object.__setattr__(self, 'save_path', save_path)
        object.__setattr__(self, 'header_written', False)  # tracks if CSV header exists
        object.__setattr__(self, 'batch_count', 0)         # counts batches processed

    def corrupt_batch(self, positive_batch: torch.LongTensor) -> torch.LongTensor:

        # Generate negatives using BasicNegativeSampler's random relation corruption
        negatives = super().corrupt_batch(positive_batch=positive_batch)

        # Increment batch counter using object.__setattr__ (same reason as __init__)
        object.__setattr__(self, 'batch_count', self.batch_count + 1)

        # Debug messages — confirm corrupt_batch is being called
        if self.batch_count == 1:
            print(f"[DEBUG] corrupt_batch called for the first time!")
            print(f"[DEBUG] Negative shape: {negatives.shape}")
        if self.batch_count % 100 == 0:
            print(f"[DEBUG] Batches captured so far: {self.batch_count}")

        try:
            # Convert integer IDs back to readable phrase labels
            neg_flat = negatives.detach().cpu().reshape(-1, 3)
            rows = []
            for t in neg_flat:
                try:
                    rows.append({
                        "head":     id_to_entity.get(t[0].item(), "unknown"),
                        "relation": id_to_relation.get(t[1].item(), "unknown"),
                        "tail":     id_to_entity.get(t[2].item(), "unknown"),
                        "label":    0,   
                    })
                except Exception as row_err:
                    print(f"[WARN] Skipping one row in batch {self.batch_count}: {row_err}")
                    continue
            if rows:
                df_batch = pd.DataFrame(rows)
                df_batch.to_csv(
                    self.save_path,
                    mode="w" if not self.header_written else "a",
                    header=not self.header_written,
                    index=False,
                )
                object.__setattr__(self, 'header_written', True)
                del df_batch
            del neg_flat, rows
        except Exception as batch_err:
            print(f"[WARN] Skipping batch {self.batch_count} due to error: {batch_err}")
        return negatives


# ══════════════════════════════════════════════════════════════════════════════
# Relation-rank evaluator
# ══════════════════════════════════════════════════════════════════════════════
class RelationRankEvaluator:

    def __init__(self, model, triples_factory, batch_size=256, device=None):
        self.model         = model
        self.num_relations = triples_factory.num_relations
        self.batch_size    = batch_size
        self.device        = device or ("cuda" if torch.cuda.is_available() else "cpu")

    def evaluate(self, mapped_triples, additional_filter_triples=None):

        self.model.eval()
        self.model.to(self.device)

        # Build filter map: (h_id, t_id) → set of all known true relation IDs
        # Used to exclude other known true relations when ranking (filtered eval)
        filter_map = {}
        if additional_filter_triples:
            for triples in additional_filter_triples:
                for triple in triples:
                    h, r, t = triple[0].item(), triple[1].item(), triple[2].item()
                    filter_map.setdefault((h, t), set()).add(r)

        eval_rows = []
        print(f"  Evaluating {len(mapped_triples):,} test triples...")

        with torch.no_grad():
            for i in range(0, len(mapped_triples), self.batch_size):
                batch = mapped_triples[i : i + self.batch_size]

                for triple in batch:
                    h_id = triple[0].item()
                    r_id = triple[1].item()
                    t_id = triple[2].item()

                    # Score all possible relations for this (h, ?, t) query
                    all_rels = torch.arange(self.num_relations, device=self.device)
                    h_tensor = torch.full((self.num_relations,), h_id, device=self.device)
                    t_tensor = torch.full((self.num_relations,), t_id, device=self.device)
                    hrt_all  = torch.stack([h_tensor, all_rels, t_tensor], dim=1)
                    scores   = self.model.score_hrt(hrt_all).squeeze()

                    masked_scores = scores.clone()
                    known_rels    = filter_map.get((h_id, t_id), set())
                    for r_other in known_rels:
                        if r_other != r_id:
                            masked_scores[r_other] = float('-inf')

                    # Rank = number of relations scoring strictly higher + 1
                    true_score = masked_scores[r_id].item()
                    rank       = int((masked_scores > true_score).sum().item()) + 1

                    eval_rows.append({
                        "head":            id_to_entity.get(h_id, "unknown"),
                        "relation":        id_to_relation.get(r_id, "unknown"),
                        "tail":            id_to_entity.get(t_id, "unknown"),
                        "relation_rank":   rank,
                        "reciprocal_rank": round(1.0 / rank, 6),
                        "hits_at_1":       int(rank <= 1),
                        "hits_at_3":       int(rank <= 3),
                        "hits_at_10":      int(rank <= 10),
                    })

                # Progress update every 20 batches
                if (i // self.batch_size + 1) % 20 == 0:
                    done = min(i + self.batch_size, len(mapped_triples))
                    print(f"  Processed {done:,} / {len(mapped_triples):,}")

        # Compute aggregate metrics
        n       = len(eval_rows)
        mrr     = sum(r["reciprocal_rank"] for r in eval_rows) / n
        h_at_1  = sum(r["hits_at_1"]       for r in eval_rows) / n
        h_at_3  = sum(r["hits_at_3"]       for r in eval_rows) / n
        h_at_10 = sum(r["hits_at_10"]      for r in eval_rows) / n
        mr      = sum(r["relation_rank"]   for r in eval_rows) / n


        metrics_dict = {
            "both": {
                "realistic": {
                    "inverse_harmonic_mean_rank": round(mrr,     4),  # MRR
                    "arithmetic_mean_rank":       round(mr,      4),  # Mean Rank
                    "hits_at_1":                  round(h_at_1,  4),
                    "hits_at_3":                  round(h_at_3,  4),
                    "hits_at_10":                 round(h_at_10, 4),
                }
            }
        }

        return metrics_dict, pd.DataFrame(eval_rows)


# ── Experiment recorder ───────────────────────────────────────────────────────
def record_results(filepath, settings, metrics, elapsed, eval_label="PyKEEN"):

    headers = [
        "No.", "Timestamp", "Dataset", "Model", "Loss Function",
        "Neg Sampler", "Corruption", "Embedding Dim", "Num Epochs",
        "Best Epoch", "Batch Size", "Learning Rate", "Margin",
        "Adv Temp", "Num Neg", "Training Time", "Evaluator",
        "MRR", "Hits@1", "Hits@3", "Hits@10", "Mean Rank",
    ]

    header_fill  = PatternFill("solid", fgColor="0D1B4B")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Load existing file or create new one
    if Path(filepath).exists():
        wb       = load_workbook(filepath)
        ws       = wb.active
        next_row = ws.max_row + 1
        run_num  = next_row - 1   # row 1 is the header
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Experiment Results"
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"    # keep header visible when scrolling
        next_row = 2
        run_num  = 1

    # Build the row with all settings and metrics
    row_data = [
        run_num,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        settings["dataset"],
        settings["model"],
        settings["loss_func"],
        settings["neg_sampler"],
        settings["corrupt"],
        settings["embedding_dim"],
        settings["num_epochs"],
        settings.get("best_epoch", "—"),
        settings["batch_size"],
        settings["learning_rate"],
        settings["margin"],
        settings["adv_temp"],
        settings["num_neg"],
        f"{int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s",
        eval_label,
        round(metrics["mrr"],        4),
        round(metrics["hits_at_1"],  4),
        round(metrics["hits_at_3"],  4),
        round(metrics["hits_at_10"], 4),
        round(metrics["mean_rank"],  4),
    ]

    for col, val in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col, value=val)

    wb.save(filepath)
    print(f"Recorded: {filepath}  Run #{run_num}")


# ── Stratified split ──────────────────────────────────────────────────────────
def stratified_split(df, train_ratio=0.8, val_ratio=0.1, random_seed=42):

    train_dfs, val_dfs, test_dfs = [], [], []

    for (domain, relation), group in df.groupby(['domain', 'relation']):
        n = len(group)

        # Groups too small to split — add entirely to train
        if n < 10:
            print(f"  [WARN] {domain} | {relation} — only {n} rows, added to train")
            train_dfs.append(group)
            continue

        # First split off test set
        train_val, test = train_test_split(
            group,
            test_size=1 - train_ratio - val_ratio,
            random_state=random_seed,
        )

        # Then split validation from remaining train+val
        val_size = val_ratio / (train_ratio + val_ratio)
        train, val = train_test_split(
            train_val,
            test_size=val_size,
            random_state=random_seed,
        )

        train_dfs.append(train)
        val_dfs.append(val)
        test_dfs.append(test)

        print(f"  {domain:12} | {relation:15} — "
              f"train: {len(train):,}  val: {len(val):,}  test: {len(test):,}")

    # Combine all groups and shuffle so domains/relations are mixed
    df_train = pd.concat(train_dfs).sample(frac=1, random_state=random_seed).reset_index(drop=True)
    df_val   = pd.concat(val_dfs  ).sample(frac=1, random_state=random_seed).reset_index(drop=True)
    df_test  = pd.concat(test_dfs ).sample(frac=1, random_state=random_seed).reset_index(drop=True)

    return df_train, df_val, df_test


# ── ABA node level classifier ─────────────────────────────────────────────────
def get_node_level(entity_name):
    """
    Returns the ABA hierarchy level for a given entity name.
        'claim'    — one of the 8 fixed claim roots
        'attacker' — any entity starting with 'no_evident_'
        'body'     — everything else (assumptions/bodies)
    """
    if entity_name in CLAIMS:
        return 'claim'          # Level 1 — root
    elif entity_name.startswith('no_evident_'):
        return 'attacker'       # Level 3 — leaf attackers
    else:
        return 'body'           # Level 2 — assumptions/bodies


# ── ABA tree builder ──────────────────────────────────────────────────────────
def build_aba_tree(df_orig, claims):
    """
    Builds the ABA tree structure from the scored triples DataFrame.

    """
    tree_rows = []

    for claim in claims:
        # Level 1→2: find all bodies that SUPPORT this claim
        bodies = df_orig[
            (df_orig['tail'] == claim) &
            (df_orig['relation'] == 'SUPPORT')
        ]

        for _, b_row in bodies.iterrows():
            body       = b_row['head']
            score_body = b_row['score']

            # Level 2→3: find all attackers of this body
            attackers = df_orig[
                (df_orig['tail'] == body) &
                (df_orig['relation'].isin(['CONTRARY_TO', 'NOT_CONTRARY']))
            ]

            if len(attackers) == 0:
                # Body has no attackers — still record it as a leaf body
                tree_rows.append({
                    'claim':          claim,
                    'body':           body,
                    'attacker':       None,
                    'body_relation':  'SUPPORT',
                    'attack_relation':None,
                    'body_score':     round(score_body, 6),
                    'attack_score':   None,
                    'domain':         b_row.get('domain', ''),
                    'body_level':     get_node_level(body),
                })
            else:
                for _, a_row in attackers.iterrows():
                    tree_rows.append({
                        'claim':           claim,
                        'body':            body,
                        'attacker':        a_row['head'],
                        'body_relation':   'SUPPORT',
                        'attack_relation': a_row['relation'],
                        'body_score':      round(score_body, 6),
                        'attack_score':    round(a_row['score'], 6),
                        'domain':          a_row.get('domain', ''),
                        'body_level':      get_node_level(body),
                    })

    return pd.DataFrame(tree_rows)


# ══════════════════════════════════════════════════════════════════════════════
# Main training pipeline
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":

    # ── Load data ─────────────────────────────────────────────────────────────
    print("=" * 60)
    print("Loading triples")
    print("=" * 60)

    df = pd.read_csv(CSV_PATH)
    df_raw = df.copy()
    df = df.drop_duplicates(subset=['head', 'relation', 'tail']).reset_index(drop=True)
    print(f"Raw rows   : {len(df_raw):,}")
    print(f"Clean rows : {len(df):,}")
    print(f"Dropped    : {len(df_raw) - len(df):,}")
    print(f"Relations found: {df['relation'].value_counts().to_dict()}")

    # Extract only the 3 columns PyKEEN needs — head, relation, tail
    triples_array = df[['head', 'relation', 'tail']].values.astype(str)

    # ── Stratified split and build TriplesFactory ─────────────────────────────
    print("=" * 60)
    print("Stratified split by domain + relation")
    print("=" * 60)

    df_train, df_val, df_test = stratified_split(
        df,
        train_ratio=TRAIN_RATIO,
        val_ratio=VAL_RATIO,
        random_seed=RANDOM_SEED,
    )

    print(f"\nTrain : {len(df_train):,}")
    print(f"Val   : {len(df_val):,}")
    print(f"Test  : {len(df_test):,}")

    # Verify domain × relation balance in test set
    print("\nDomain × Relation distribution in test set:")
    print(df_test.groupby(['domain', 'relation']).size().to_string())

    # Build TriplesFactory from stratified train split
    # validation and testing MUST share the same entity/relation vocabulary
    # as training — otherwise IDs would not match and embeddings would be wrong
    training = TriplesFactory.from_labeled_triples(
        triples=df_train[['head', 'relation', 'tail']].values.astype(str),
    )
    validation = TriplesFactory.from_labeled_triples(
        triples=df_val[['head', 'relation', 'tail']].values.astype(str),
        entity_to_id=training.entity_to_id,       # share training vocabulary
        relation_to_id=training.relation_to_id,   # share training vocabulary
    )
    testing = TriplesFactory.from_labeled_triples(
        triples=df_test[['head', 'relation', 'tail']].values.astype(str),
        entity_to_id=training.entity_to_id,       # share training vocabulary
        relation_to_id=training.relation_to_id,   # share training vocabulary
    )

    print(f"\nUnique entities  : {training.num_entities:,}")
    print(f"Unique relations : {training.num_relations}")

    # Build lookup dictionaries — used inside ExtendedBasicNegativeSampler
    # and RelationRankEvaluator to convert IDs back to readable labels
    id_to_entity   = {v: k for k, v in training.entity_to_id.items()}
    id_to_relation = {v: k for k, v in training.relation_to_id.items()}

    # ── Train ─────────────────────────────────────────────────────────────────
    print("=" * 60)
    print("Training " + MODEL)
    print("=" * 60)

    start_time = time.time()

    result = pipeline(
        training=training,
        validation=validation,
        testing=testing,

        model=MODEL,
        model_kwargs=dict(embedding_dim=EMBEDDING_DIM),

        loss=LOSS,
        loss_kwargs=dict(
            margin=MARGIN,
            adversarial_temperature=ADV_TEMP,
        ),

        negative_sampler=ExtendedBasicNegativeSampler,
        negative_sampler_kwargs=dict(
            num_negs_per_pos=NUM_NEG,
            corruption_scheme=[CORRUPT],   
        ),


        optimizer="Adam",
        optimizer_kwargs=dict(lr=LEARNING_RATE),

        training_loop="sLCWA",
        training_kwargs=dict(
            num_epochs=NUM_EPOCHS,
            batch_size=BATCH_SIZE,
        ),


        stopper="early",
        stopper_kwargs=dict(
            metric="hits_at_10",
            patience=10,
            relative_delta=0.002,
            frequency=10,
        ),

        evaluator="RankBasedEvaluator",
        evaluator_kwargs=dict(filtered=True),

        random_seed=RANDOM_SEED,

        device="cuda" if torch.cuda.is_available() else "cpu",
    )

    end_time = time.time()
    elapsed  = end_time - start_time
    print(f"\nTraining completed in: {int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s")

    # ── PyKEEN default evaluation results ─────────────────────────────────────
    print("=" * 60)
    print("Evaluation Results (PyKEEN default)")
    print("=" * 60)

    metrics_dict = result.metric_results.to_dict()
    r_default    = metrics_dict["both"]["realistic"]

    mrr        = r_default['inverse_harmonic_mean_rank']
    hits_at_1  = r_default['hits_at_1']
    hits_at_3  = r_default['hits_at_3']
    hits_at_10 = r_default['hits_at_10']
    mean_rank  = r_default['arithmetic_mean_rank']

    print(f"MRR       : {mrr:.4f}")
    print(f"Hits@1    : {hits_at_1:.4f}")
    print(f"Hits@3    : {hits_at_3:.4f}")
    print(f"Hits@10   : {hits_at_10:.4f}")
    print(f"Mean Rank : {mean_rank:.4f}")

    # ── Record results to Excel ───────────────────────────────────────────────
    try:
        best_epoch = result.stopper.best_epoch
    except Exception:
        best_epoch = "—"

    settings = {
        "dataset":       DATASET,
        "model":         MODEL,
        "loss_func":     LOSS,
        "neg_sampler":   "ExtendedBasicNegativeSampler",
        "corrupt":       CORRUPT,
        "embedding_dim": EMBEDDING_DIM,
        "num_epochs":    NUM_EPOCHS,
        "best_epoch":    best_epoch,
        "batch_size":    BATCH_SIZE,
        "learning_rate": LEARNING_RATE,
        "margin":        MARGIN,
        "adv_temp":      ADV_TEMP,
        "num_neg":       NUM_NEG,
    }

    metrics_out = {
        "mrr":        mrr,
        "hits_at_1":  hits_at_1,
        "hits_at_3":  hits_at_3,
        "hits_at_10": hits_at_10,
        "mean_rank":  mean_rank,
    }

    record_results(RESULTS_FILE, settings, metrics_out, elapsed, eval_label="PyKEEN")

    # ── Relation-rank evaluation ──────────────────────────────────────────────
    print("=" * 60)
    print("Relation-Rank Evaluation (aligned with relation corruption)")
    print("=" * 60)

    try:
        rel_evaluator = RelationRankEvaluator(
            model=result.model,
            triples_factory=training,
            batch_size=BATCH_SIZE,
            device="cuda" if torch.cuda.is_available() else "cpu",
        )

        # evaluate() returns same nested dict format as PyKEEN's .to_dict()
        rel_metrics_dict, df_rel_eval = rel_evaluator.evaluate(
            mapped_triples=testing.mapped_triples,
            additional_filter_triples=[
                training.mapped_triples,
                validation.mapped_triples,
            ],
        )

        # Access using same pattern as PyKEEN default evaluation above
        r_rel = rel_metrics_dict["both"]["realistic"]

        rel_mrr = r_rel["inverse_harmonic_mean_rank"]
        rel_h1  = r_rel["hits_at_1"]
        rel_h3  = r_rel["hits_at_3"]
        rel_h10 = r_rel["hits_at_10"]
        rel_mr  = r_rel["arithmetic_mean_rank"]

        # Print side-by-side comparison — relation evaluator vs PyKEEN default
        print(f"\n{'Metric':<12} {'RelationRank':>14}  {'PyKEEN (entity)':>16}")
        print("-" * 46)
        print(f"{'MRR':<12} {rel_mrr:>14.4f}  {mrr:>16.4f}")
        print(f"{'Hits@1':<12} {rel_h1:>14.4f}  {hits_at_1:>16.4f}")
        print(f"{'Hits@3':<12} {rel_h3:>14.4f}  {hits_at_3:>16.4f}")
        print(f"{'Hits@10':<12} {rel_h10:>14.4f}  {hits_at_10:>16.4f}")
        print(f"{'Mean Rank':<12} {rel_mr:>14.4f}  {mean_rank:>16.4f}")

        # Record relation-rank results to Excel as a separate row
        metrics_rel_out = {
            "mrr":        rel_mrr,
            "hits_at_1":  rel_h1,
            "hits_at_3":  rel_h3,
            "hits_at_10": rel_h10,
            "mean_rank":  rel_mr,
        }
        record_results(
            RESULTS_FILE, settings, metrics_rel_out, elapsed,
            eval_label="RelationRank"
        )
        #drop duplicate and add domain in eval results
        df_orig_for_join = pd.read_csv(CSV_PATH)[['head', 'relation', 'tail', 'domain']].drop_duplicates(
            subset=['head', 'relation', 'tail']
        )
        df_rel_eval = df_rel_eval.merge(
            df_orig_for_join,
            on=['head', 'relation', 'tail'],
            how='left'
        )
        # Save per-triple relation ranks — overwrites each run
        rel_eval_path = OUTPUT_DIR / "relation_eval_ranked.csv"
        df_rel_eval.to_csv(rel_eval_path, index=False, mode='w')
        print(f"\nSaved: {rel_eval_path}  ({len(df_rel_eval):,} triples)")

    except Exception as e:
        print(f"[ERROR] Relation-rank evaluation failed: {e}")
        import traceback; traceback.print_exc()

    # ── Export visualization CSV + ABA tree structure ─────────────────────────
    print("=" * 60)
    print("Exporting visualization data + ABA tree structure")
    print("=" * 60)

    try:
        # Extract RotatE entity embeddings
        entity_emb_complex = result.model.entity_representations[0](indices=None).detach().cpu()
        entity_emb_np      = entity_emb_complex.numpy()
        entity_emb_real    = np.concatenate([
            entity_emb_np.real,
            entity_emb_np.imag,
        ], axis=1).astype(np.float32)

        print(f"Entity embedding shape (real+imag): {entity_emb_real.shape}")

        # Project to 2D with PCA
        pca       = PCA(n_components=2, random_state=RANDOM_SEED)
        coords_2d = pca.fit_transform(entity_emb_real)
        print(f"PCA explained variance ratio: {pca.explained_variance_ratio_.round(4)}")

        # Build entity dataframe with 2D coordinates
        entity_to_id = training.entity_to_id
        entity_rows  = []
        for entity_name, entity_id in entity_to_id.items():
            row_data={
                "entity_id":   entity_id,
                "entity_name": entity_name,
                "x":           round(float(coords_2d[entity_id, 0]), 6),
                "y":           round(float(coords_2d[entity_id, 1]), 6),
            }
            for dim_idx, val in enumerate(entity_emb_real[entity_id]):
                row_data[f"emb_{dim_idx}"] = round(float(val), 6)
    
            entity_rows.append(row_data) 
        df_entities = pd.DataFrame(entity_rows)

        # Score all original triples
        df_orig     = pd.read_csv(CSV_PATH)  # keep ALL columns including domain
        scores_list = []
        result.model.eval()
        with torch.no_grad():
            for _, row in df_orig.iterrows():
                try:
                    h_id = training.entity_to_id.get(str(row['head']), -1)
                    r_id = training.relation_to_id.get(str(row['relation']), -1)
                    t_id = training.entity_to_id.get(str(row['tail']), -1)
                    if -1 in (h_id, r_id, t_id):
                        scores_list.append(None)
                        continue
                    hrt = torch.tensor([[h_id, r_id, t_id]], dtype=torch.long)
                    s   = result.model.score_hrt(hrt)
                    scores_list.append(round(float(s.item()), 6))
                except Exception:
                    scores_list.append(None)

        df_orig['score'] = scores_list

        # Join 2D coordinates
        coord_map         = df_entities.set_index('entity_name')[['x', 'y']]
        df_orig['head_x'] = df_orig['head'].map(coord_map['x'])
        df_orig['head_y'] = df_orig['head'].map(coord_map['y'])
        df_orig['tail_x'] = df_orig['tail'].map(coord_map['x'])
        df_orig['tail_y'] = df_orig['tail'].map(coord_map['y'])

        # Drop rows with missing coords or scores
        df_orig = df_orig.dropna(subset=['score', 'head_x', 'head_y', 'tail_x', 'tail_y'])

        # Save standard visualization CSV (for Graph View)
        VIZ_FILE = OUTPUT_DIR / "visualization_data.csv"
        df_orig[['head', 'relation', 'tail', 'score',
                 'head_x', 'head_y', 'tail_x', 'tail_y']].to_csv(VIZ_FILE, index=False)
        print(f"Saved: {VIZ_FILE}  ({len(df_orig):,} triples)")

        # Build and save ABA tree structure (now uses the module-level function)
        df_tree   = build_aba_tree(df_orig, CLAIMS)
        TREE_FILE = OUTPUT_DIR / "aba_tree_structure.csv"
        df_tree.to_csv(TREE_FILE, index=False)
        print(f"Saved: {TREE_FILE}  ({len(df_tree):,} rows, {df_tree['claim'].nunique()} claims)")
        print(f"  Claims:   {df_tree['claim'].nunique()}")
        print(f"  Bodies:   {df_tree['body'].nunique()}")
        print(f"  Attackers:{df_tree['attacker'].nunique()}")

        # Save entity embeddings
        ENTITY_FILE = OUTPUT_DIR / "entity_embeddings.csv"
        df_entities.to_csv(ENTITY_FILE, index=False)
        print(f"Saved: {ENTITY_FILE}  ({len(df_entities):,} entities)")

    except Exception as e:
        print(f"[ERROR] Visualization export failed: {e}")
        import traceback; traceback.print_exc()

    model_save_path = OUTPUT_DIR / f"trained_model_{MODEL}.pkl"
    tf_save_path    = OUTPUT_DIR / f"triples_factory_{MODEL}.pkl"
    
    # Save trained model weights
    with open(model_save_path, "wb") as f:
        pickle.dump(result.model, f)
    print(f"Saved model to {model_save_path}")
    
    # Save TriplesFactory (entity/relation ID mappings)
    with open(tf_save_path, "wb") as f:
        pickle.dump(training, f)
    print(f"Saved TF to {tf_save_path}")
    df_test[['head','relation','tail']].to_csv(OUTPUT_DIR / "test_triples.csv", index=False)

    print("=" * 60)
    print("All done.")
    print("=" * 60)