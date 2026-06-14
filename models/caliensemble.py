"""
Calibrated KGE Ensemble for ABA Relation Prediction
=====================================================
Claude-assisted

"""

import pickle
import datetime
import itertools
from pathlib import Path

import numpy as np
import pandas as pd
import torch
from sklearn.calibration import CalibratedClassifierCV
from sklearn.linear_model import LogisticRegression
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════

ROOT = Path(__file__).resolve().parent.parent
# pkl file paths of each KGE model
MODEL_PATHS = {
    "RotatE":   ROOT / "outputs" / "rotate_sup_output"  / "trained_model_RotatE.pkl",
    "ComplEx":  ROOT / "outputs" / "complex_sup_output" / "trained_model_ComplEx.pkl",
    "TransE":   ROOT / "outputs" / "transe_sup_output"  / "trained_model_TransE.pkl",
    "DistMult": ROOT / "outputs" / "distmult_sup_output"/ "trained_model_DistMult.pkl",
    "ConvKB":   ROOT / "outputs" / "convkb_sup_output"  / "trained_model_ConvKB.pkl",
}

TF_PATH = ROOT / "outputs" / "rotate_sup_output" / "triples_factory_RotatE.pkl"
TEST_CSV = ROOT / "outputs" / "rotate_sup_output" / "test_triples.csv"
RESULTS_FILE = ROOT / "data" / "experiment_results" / "sup_experiment_results.xlsx"

# Output folder for per-prediction CSVs
OUTPUT_DIR = ROOT / "outputs" / "ensemble_output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Calibration settings (matching the paper)
CV_FOLDS  = 5      # 5-fold cross-validation for calibrator fitting
MAX_ITER  = 1000   # logistic regression max iterations
BATCH_SIZE = 512   # GPU batch size for scoring

# Ensemble experiment ID
ENSEMBLE_EXP_ID = f"exp_ensemble_{datetime.datetime.now().strftime('%Y%m%d')}_3R"

# ══════════════════════════════════════════════════════════════════════════════
# Load test triples from CSV using TriplesFactory mapping
# ══════════════════════════════════════════════════════════════════════════════

def load_test_triples(csv_path, tf):
    """
    Reads test_triples.csv and maps entity/relation strings to integer IDs.
    Returns three numpy arrays: heads, tails, true_rel_ids.
    """
    df = pd.read_csv(csv_path)
    df.columns = [c.strip().lower() for c in df.columns]

    entity2id   = tf.entity_to_id
    relation2id = tf.relation_to_id

    heads, tails, rels = [], [], []
    skipped = 0
    for _, row in df.iterrows():
        h = str(row["head"]).strip()
        r = str(row["relation"]).strip()
        t = str(row["tail"]).strip()
        if h not in entity2id or t not in entity2id or r not in relation2id:
            skipped += 1
            continue
        heads.append(entity2id[h])
        tails.append(entity2id[t])
        rels.append(relation2id[r])

    if skipped:
        print(f"  Warning: {skipped} test triples skipped (unseen entities/relations)")

    return np.array(heads), np.array(tails), np.array(rels)


# ══════════════════════════════════════════════════════════════════════════════
# Score all relations for every (h, t) pair
# ══════════════════════════════════════════════════════════════════════════════

def get_relation_scores(model, num_relations, heads, tails, device):
    """
    For each (h, t) pair, compute model scores for ALL relations.
    Returns array of shape (N, num_relations).

    """
    model.eval()
    model.to(device)

    all_scores = []
    h_tensor = torch.tensor(heads, dtype=torch.long, device=device)
    t_tensor = torch.tensor(tails, dtype=torch.long, device=device)

    with torch.no_grad():
        for start in range(0, len(heads), BATCH_SIZE):
            end = min(start + BATCH_SIZE, len(heads))
            h_batch = h_tensor[start:end]
            t_batch = t_tensor[start:end]

            batch_rel_scores = []
            for r_id in range(num_relations):
                r_batch = torch.full((end - start,), r_id,
                                     dtype=torch.long, device=device)
                hrt = torch.stack([h_batch, r_batch, t_batch], dim=1)
                scores = model.score_hrt(hrt).squeeze(-1).cpu().numpy()
                batch_rel_scores.append(scores)

            # shape: (num_relations, B) → transpose → (B, num_relations)
            all_scores.append(np.stack(batch_rel_scores, axis=0).T)

    return np.vstack(all_scores)  # (N, num_relations)


# ══════════════════════════════════════════════════════════════════════════════
# Calibrate raw scores to probabilities
# logistic sigmoid per relation (one-vs-rest) + 5-fold CV
# ══════════════════════════════════════════════════════════════════════════════

def load_train_triples(tf, test_csv_path):
    """
    Extract training triples from TriplesFactory by subtracting the saved test set.
    Returns (train_heads, train_tails, train_rel_ids) as numpy arrays.
    """
    # Build a set of test (h, r, t) ID tuples to exclude
    test_df = pd.read_csv(test_csv_path)
    test_df.columns = [c.strip().lower() for c in test_df.columns]
    e2id = tf.entity_to_id
    r2id = tf.relation_to_id
    test_set = set()
    for _, row in test_df.iterrows():
        h = e2id.get(str(row["head"]).strip())
        r = r2id.get(str(row["relation"]).strip())
        t = e2id.get(str(row["tail"]).strip())
        if h is not None and r is not None and t is not None:
            test_set.add((int(h), int(r), int(t)))

    # All triples in the factory
    all_triples = tf.mapped_triples.numpy()   # (N_total, 3)
    mask = np.array(
        [(int(h), int(r), int(t)) not in test_set for h, r, t in all_triples]
    )
    train = all_triples[mask]
    # mapped_triples columns are (h, r, t) at indices [0, 1, 2].
    return train[:, 0], train[:, 2], train[:, 1]   # heads, tails, rel_ids


def fit_calibrators(raw_train_scores, train_rel_ids, num_relations):
    """
    Fit one binary logistic calibrator per relation on TRAINING scores.
    Returns a list of fitted calibrators
    """
    calibrators = []
    for r in range(num_relations):
        binary_labels = (train_rel_ids == r).astype(int)
        feature = raw_train_scores[:, r].reshape(-1, 1)
        if len(np.unique(binary_labels)) < 2:
            calibrators.append(None)
            continue
        base_clf = LogisticRegression(max_iter=MAX_ITER, solver="lbfgs", class_weight="balanced")
        cal = CalibratedClassifierCV(estimator=base_clf, method="sigmoid", cv=CV_FOLDS)
        cal.fit(feature, binary_labels)
        calibrators.append(cal)
    return calibrators


def apply_calibrators(calibrators, raw_test_scores, num_relations):
    """
    Apply pre-fitted calibrators to TEST scores to normalized probability array.
    """
    N = raw_test_scores.shape[0]
    calibrated = np.zeros((N, num_relations), dtype=np.float64)
    for r, cal in enumerate(calibrators):
        feature = raw_test_scores[:, r].reshape(-1, 1)
        if cal is None:
            calibrated[:, r] = 1.0 / (1.0 + np.exp(-feature.squeeze()))
        else:
            calibrated[:, r] = cal.predict_proba(feature)[:, 1]
    row_sums = calibrated.sum(axis=1, keepdims=True)
    row_sums = np.where(row_sums == 0, 1.0, row_sums)
    return calibrated / row_sums

# ══════════════════════════════════════════════════════════════════════════════
# Compute relation-rank metrics
# ══════════════════════════════════════════════════════════════════════════════

def compute_metrics(probs, true_rel_ids, relation_id2name):
    """
    Rank relations by descending probability for each test triple.
    Compute MRR, MR, micro/macro Hits@1, Hits@3/10 overall and per relation.

    Micro Hits@1 : fraction of all test triples ranked correctly at position 1.
    Macro Hits@1 : mean of per-relation Hits@1 (equal weight per relation class).
    """
    N = len(true_rel_ids)
    ranks = np.zeros(N, dtype=int)

    for i in range(N):
        sorted_rels = np.argsort(-probs[i])          # descending probability
        rank = np.where(sorted_rels == true_rel_ids[i])[0][0] + 1  # 1-indexed
        ranks[i] = rank

    def mrr():   return float((1.0 / ranks).mean())
    def h_at(k): return float((ranks <= k).mean())

    per_rel = {}
    for r_id, r_name in relation_id2name.items():
        mask = true_rel_ids == r_id
        if not mask.any():
            continue
        r_ranks = ranks[mask]
        per_rel[r_name] = {
            "MRR":     float((1.0 / r_ranks).mean()),
            "Hits@1":  float((r_ranks <= 1).mean()),
            "Hits@3":  float((r_ranks <= 3).mean()),
            "Hits@10": float((r_ranks <= 10).mean()),
            "MeanRank": float(r_ranks.mean()),
            "N": int(mask.sum()),
        }

    macro_h1 = float(np.mean([v["Hits@1"] for v in per_rel.values()])) if per_rel else 0.0

    overall = {
        "MRR":       mrr(),
        "MeanRank":  float(ranks.mean()),
        "MicroH@1":  h_at(1),
        "MacroH@1":  macro_h1,
        "Hits@3":    h_at(3),
        "Hits@10":   h_at(10),
        "N": N,
    }

    return {"overall": overall, "per_relation": per_rel}


# ══════════════════════════════════════════════════════════════════════════════
# Record to Excel
# ══════════════════════════════════════════════════════════════════════════════

def record_ensemble_to_excel(filepath, all_results, timestamp):

    if Path(filepath).exists():
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    header_fill  = PatternFill("solid", fgColor="0D1B4B")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def make_or_get_sheet(wb, name, headers):
        if name not in wb.sheetnames:
            ws = wb.create_sheet(name)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            ws.freeze_panes = "A2"
        return wb[name]

    t2_headers = [
        "timestamp", "ensemble_exp_id", "model_combo", "type",
        "MRR", "MeanRank", "MicroHits@1", "MacroHits@1", "Hits@3", "Hits@10", "N_test",
        "CONTRARY_TO_MRR", "CONTRARY_TO_H1",
        "NOT_CONTRARY_MRR", "NOT_CONTRARY_H1",
        "SUPPORT_MRR", "SUPPORT_H1",
    ]
    ws2 = make_or_get_sheet(wb, "Ensemble_Table2", t2_headers)

    for entry in all_results:
        pr = entry["metrics"]["per_relation"]
        row = [
            timestamp,
            entry["exp_id"],
            entry["combo_label"],
            "individual" if len(entry["models"]) == 1 else f"ensemble_{len(entry['models'])}",
            round(entry["metrics"]["overall"]["MRR"],      4),
            round(entry["metrics"]["overall"]["MeanRank"], 4),
            round(entry["metrics"]["overall"]["MicroH@1"], 4),
            round(entry["metrics"]["overall"]["MacroH@1"], 4),
            round(entry["metrics"]["overall"]["Hits@3"],   4),
            round(entry["metrics"]["overall"]["Hits@10"],  4),
            entry["metrics"]["overall"]["N"],
            round(pr.get("CONTRARY_TO",  {}).get("MRR",    0), 4),
            round(pr.get("CONTRARY_TO",  {}).get("Hits@1", 0), 4),
            round(pr.get("NOT_CONTRARY", {}).get("MRR",    0), 4),
            round(pr.get("NOT_CONTRARY", {}).get("Hits@1", 0), 4),
            round(pr.get("SUPPORT",      {}).get("MRR",    0), 4),
            round(pr.get("SUPPORT",      {}).get("Hits@1", 0), 4),
        ]
        ws2.append(row)

    t3_headers = [
        "timestamp", "ensemble_exp_id", "ensemble_combo", "vs_model",
        "pct_change_MRR", "pct_change_MR (neg=better)",
        "pct_change_MicroHits@1", "pct_change_MacroHits@1",
        "pct_change_Hits@3", "pct_change_Hits@10",
    ]
    ws3 = make_or_get_sheet(wb, "Ensemble_Table3", t3_headers)

    # Index individual model results for lookup
    indiv_lookup = {
        e["combo_label"]: e["metrics"]["overall"]
        for e in all_results if len(e["models"]) == 1
    }

    for entry in all_results:
        if len(entry["models"]) <= 1:
            continue
        ens_metrics = entry["metrics"]["overall"]
        for model_name in entry["models"]:
            if model_name not in indiv_lookup:
                continue
            base = indiv_lookup[model_name]
            def pct(metric):
                b = base[metric]
                if b == 0:
                    return None
                return round(100 * (ens_metrics[metric] - b) / b, 1)
            ws3.append([
                timestamp,
                entry["exp_id"],
                entry["combo_label"],
                model_name,
                pct("MRR"),
                pct("MeanRank"),
                pct("MicroH@1"),
                pct("MacroH@1"),
                pct("Hits@3"),
                pct("Hits@10"),
            ])

    wb.save(filepath)
    print(f"\n  Results recorded → {filepath}")
    print(f"  Sheets: Ensemble_Table2, Ensemble_Table3")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"\nDevice: {device}")

    # ── Filter to only models whose .pkl files exist ──────────────────────────
    available_models = {
        name: path for name, path in MODEL_PATHS.items()
        if Path(path).exists()
    }
    if not available_models:
        raise FileNotFoundError(
            "No trained model .pkl files found. "
            "Add the save_model_snippet.py block to your training scripts first."
        )
    print(f"\nAvailable models: {list(available_models.keys())}")

    # ── Load TriplesFactory ───────────────────────────────────────────────────
    print("\nLoading TriplesFactory...")
    with open(TF_PATH, "rb") as f:
        tf = pickle.load(f)
    relation_id2name = {v: k for k, v in tf.relation_to_id.items()}
    num_relations    = tf.num_relations
    print(f"  Entities  : {tf.num_entities:,}")
    print(f"  Relations : {relation_id2name}")

    # ── Load test triples ─────────────────────────────────────────────────────
    print("\nLoading test triples")
    heads, tails, true_rel_ids = load_test_triples(TEST_CSV, tf)
    print(f"  Test set: {len(heads):,} triples")
    for r_id, r_name in relation_id2name.items():
        count = (true_rel_ids == r_id).sum()
        print(f"    {r_name:<20}: {count:>5} ({100*count/len(heads):.1f}%)")

    # ── Load train triples for calibration (TF minus test set) ───────────────
    print("\nExtracting train triples for calibration")
    train_heads, train_tails, train_rel_ids = load_train_triples(tf, TEST_CSV)
    print(f"  Train set: {len(train_heads):,} triples")
    for r_id, r_name in relation_id2name.items():
        count = (train_rel_ids == r_id).sum()
        print(f"    {r_name:<20}: {count:>5} ({100*count/len(train_heads):.1f}%)")

    # ── Score + calibrate each model individually ─────────────────────────────
    print("\nScoring train set, fitting calibrators, applying to test set")
    calibrated_probs = {}   # model_name → (N_test, num_relations) probability array

    for model_name, model_path in available_models.items():
        print(f"\n  ── {model_name} ──")
        with open(model_path, "rb") as f:
            model = pickle.load(f)
        model.to(device)

        print(f"     Scoring {len(train_heads):,} train pairs (for calibration)")
        raw_train_scores = get_relation_scores(model, num_relations, train_heads, train_tails, device)

        print(f"     Fitting calibrators on train set ({CV_FOLDS}-fold CV)")
        calibrators = fit_calibrators(raw_train_scores, train_rel_ids, num_relations)

        print(f"     Scoring {len(heads):,} test pairs")
        raw_test_scores = get_relation_scores(model, num_relations, heads, tails, device)

        print(f"     Applying calibrators to test scores")
        probs = apply_calibrators(calibrators, raw_test_scores, num_relations)
        calibrated_probs[model_name] = probs

        # Quick sanity check
        preds = np.argmax(probs, axis=1)
        acc = (preds == true_rel_ids).mean()
        print(f"     Calibrated accuracy: {acc:.4f}")

    # ── Try all combinations ──────────────────────────────────────────────────
    print("\nRunning all model combinations")
    model_names = list(calibrated_probs.keys())
    all_results = []

    # Individual models first
    for name in model_names:
        probs = calibrated_probs[name]
        metrics = compute_metrics(probs, true_rel_ids, relation_id2name)
        all_results.append({
            "combo_label": name,
            "models":      [name],
            "metrics":     metrics,
            "exp_id":      ENSEMBLE_EXP_ID,
        })

    # All combinations of size 2 and above
    for size in range(2, len(model_names) + 1):
        for combo in itertools.combinations(model_names, size):
            stacked = np.stack([calibrated_probs[m] for m in combo], axis=0)
            avg_probs = stacked.mean(axis=0)

            metrics = compute_metrics(avg_probs, true_rel_ids, relation_id2name)
            label   = " + ".join(combo)
            all_results.append({
                "combo_label": label,
                "models":      list(combo),
                "metrics":     metrics,
                "exp_id":      ENSEMBLE_EXP_ID,
            })
            print(f"  {label:<50} MRR={metrics['overall']['MRR']:.4f}  μH@1={metrics['overall']['MicroH@1']:.4f}  MH@1={metrics['overall']['MacroH@1']:.4f}")

    # ── Print summary ───────────────────────────────────────────
    print("\n" + "═" * 105)
    print(f" Individual and Ensemble Results  [{ENSEMBLE_EXP_ID}]")
    print("═" * 105)
    print(f"  {'Model / Combo':<45} {'MRR':>7} {'MR':>7} {'μH@1':>8} {'MH@1':>8} {'Hits@3':>8} {'Hits@10':>9}")
    print("  " + "─" * 103)
    for entry in all_results:
        m = entry["metrics"]["overall"]
        marker = "  " if len(entry["models"]) == 1 else "▶ "
        print(f"  {marker}{entry['combo_label']:<43} "
              f"{m['MRR']:>7.4f} {m['MeanRank']:>7.4f} "
              f"{m['MicroH@1']:>8.4f} {m['MacroH@1']:>8.4f} "
              f"{m['Hits@3']:>8.4f} {m['Hits@10']:>9.4f}")
    print("═" * 105)

    # ── Print percent increase ──────────────────────────────────
    indiv_lookup = {
        e["combo_label"]: e["metrics"]["overall"]
        for e in all_results if len(e["models"]) == 1
    }
    print(f"\nPercent Change vs Individual Models  (MR: negative = better)")
    print("═" * 112)
    print(f"  {'Ensemble':<35} {'vs Model':<15} {'MRR':>7} {'MR':>8} {'μH@1':>8} {'MH@1':>8} {'Hits@3':>8} {'Hits@10':>9}")
    print("  " + "─" * 110)
    for entry in all_results:
        if len(entry["models"]) <= 1:
            continue
        ens = entry["metrics"]["overall"]
        for model_name in entry["models"]:
            if model_name not in indiv_lookup:
                continue
            base = indiv_lookup[model_name]
            def pct(metric):
                b = base[metric]
                return f"{100*(ens[metric]-b)/b:+.0f}%" if b != 0 else "N/A"
            print(f"  {entry['combo_label']:<35} {model_name:<15} "
                  f"{pct('MRR'):>7} {pct('MeanRank'):>8} "
                  f"{pct('MicroH@1'):>8} {pct('MacroH@1'):>8} "
                  f"{pct('Hits@3'):>8} {pct('Hits@10'):>9}")
    print("═" * 112)

    # ── Record to Excel ───────────────────────────────────────────────────────
    print("\nRecording to Excel")
    record_ensemble_to_excel(RESULTS_FILE, all_results, timestamp)

    # ── Save results JSON ─────────────────────────────────────────────────────
    print("\nSaving results JSON")
    import json as _json

    def _pct_change(ens_val, base_val):
        if base_val and base_val != 0:
            return round(100 * (ens_val - base_val) / base_val, 4)
        return None

    json_payload = {
        "generated_at":    timestamp,
        "experiment_id":   ENSEMBLE_EXP_ID,
        "n_test_triples":  int(len(heads)),
        "available_models": list(available_models.keys()),
        "results": [],
    }

    for entry in all_results:
        ov = entry["metrics"]["overall"]
        pr = entry["metrics"]["per_relation"]

        pct_vs = {}
        if len(entry["models"]) > 1:
            for model_name in entry["models"]:
                if model_name not in indiv_lookup:
                    continue
                base = indiv_lookup[model_name]
                pct_vs[model_name] = {
                    metric: _pct_change(ov[metric], base[metric])
                    for metric in ["MRR", "MeanRank", "MicroH@1", "MacroH@1",
                                   "Hits@3", "Hits@10"]
                }

        json_payload["results"].append({
            "combo":  entry["combo_label"],
            "type":   "individual" if len(entry["models"]) == 1
                      else f"ensemble_{len(entry['models'])}",
            "models": entry["models"],
            "overall": {
                "MRR":      round(ov["MRR"],      6),
                "MR":       round(ov["MeanRank"],  6),
                "MicroH@1": round(ov["MicroH@1"], 6),
                "MacroH@1": round(ov["MacroH@1"], 6),
                "Hits@3":   round(ov["Hits@3"],   6),
                "Hits@10":  round(ov["Hits@10"],  6),
                "N":        ov["N"],
            },
            "per_relation": {
                r_name: {
                    "MRR":     round(r["MRR"],      6),
                    "MR":      round(r["MeanRank"], 6),
                    "Hits@1":  round(r["Hits@1"],   6),
                    "Hits@3":  round(r["Hits@3"],   6),
                    "Hits@10": round(r["Hits@10"],  6),
                    "N":       r["N"],
                }
                for r_name, r in pr.items()
            },
            "pct_change_vs": pct_vs,
        })

    out_json = OUTPUT_DIR / f"{ENSEMBLE_EXP_ID}_results.json"
    with open(out_json, "w") as f:
        _json.dump(json_payload, f, indent=2)
    print(f"  Saved: {out_json}")

    print("\Done.")