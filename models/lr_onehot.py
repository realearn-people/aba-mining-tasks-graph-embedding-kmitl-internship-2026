"""
One-Hot Encoding Logistic Regression Baseline
==============================================
Independently trains a Logistic Regression classifier using
one-hot encoded entity representations — no KGE embeddings needed.

Claude-Assisted
"""

import pandas as pd
import numpy as np
import datetime
import time
import warnings
from pathlib import Path

from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import OneHotEncoder, StandardScaler
from sklearn.metrics import confusion_matrix
from sklearn.model_selection import train_test_split
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings("ignore")

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT         = Path(__file__).resolve().parent.parent
RESULTS_FILE = ROOT / "data" / "experiment_results" / "sup_experiment_results.xlsx"
CSV_PATH     = ROOT / "data" / "input_data" / "hotel_contrary_dataset_support.csv"
OUTPUT_DIR   = ROOT / "outputs" / "onehot_sup_output"
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Hyperparameters ───────────────────────────────────────────────────────────
TRAIN_RATIO  = 0.8
VAL_RATIO    = 0.1
RANDOM_SEED  = 42
MAX_ITER     = 1000
C            = 1.0     

# ── ABA claims ────────────────────────────────────────────────────────────────
CLAIMS = [
    'good_staff',    'bad_staff',
    'good_price',    'bad_price',
    'good_check-in', 'bad_check-in',
    'good_check-out','bad_check-out',
]

RELATIONS = ['CONTRARY_TO', 'NOT_CONTRARY', 'SUPPORT']


# ── Stratified split (identical seed as all other scripts) ────────────────────
def stratified_split(df, train_ratio=0.8, val_ratio=0.1, random_seed=42):

    train_dfs, val_dfs, test_dfs = [], [], []

    for (domain, relation), group in df.groupby(['domain', 'relation']):
        n = len(group)
        if n < 10:
            train_dfs.append(group)
            continue

        train_val, test = train_test_split(
            group,
            test_size=1 - train_ratio - val_ratio,
            random_state=random_seed,
        )
        val_size = val_ratio / (train_ratio + val_ratio)
        train, val = train_test_split(
            train_val,
            test_size=val_size,
            random_state=random_seed,
        )
        train_dfs.append(train)
        val_dfs.append(val)
        test_dfs.append(test)

    df_train = pd.concat(train_dfs).sample(frac=1, random_state=random_seed).reset_index(drop=True)
    df_val   = pd.concat(val_dfs  ).sample(frac=1, random_state=random_seed).reset_index(drop=True)
    df_test  = pd.concat(test_dfs ).sample(frac=1, random_state=random_seed).reset_index(drop=True)

    return df_train, df_val, df_test


# ── Per-relation metrics ──────────────────────────────────────────────────────
def compute_per_relation_metrics(y_true, y_pred, relations=None):
    """
    Compute per-relation Hits@1 and Macro-Hits@1.
    Hits@1 = accuracy 
    """
    if relations is None:
        relations = sorted(set(y_true))

    rows = []
    for rel in relations:
        mask      = y_true == rel
        count     = mask.sum()
        correct   = ((y_true == rel) & (y_pred == rel)).sum()
        hits_at_1 = correct / count if count > 0 else 0.0
        rows.append({
            "relation":  rel,
            "count":     int(count),
            "correct":   int(correct),
            "hits_at_1": round(hits_at_1, 4),
        })

    df_per = pd.DataFrame(rows).set_index("relation")
    macro  = df_per["hits_at_1"].mean()
    return df_per, round(macro, 4)


# ── Experiment recorder ───────────────────────────────────────────────────────
def record_results(filepath, settings, metrics, elapsed, eval_label="OneHot-LR"):
    headers = [
        "No.", "Timestamp", "Dataset", "Model", "Loss Function",
        "Neg Sampler", "Corruption", "Embedding Dim", "Num Epochs",
        "Best Epoch", "Batch Size", "Learning Rate", "Margin",
        "Adv Temp", "Num Neg", "Training Time", "Evaluator",
        "MRR", "Hits@1", "Hits@3", "Hits@10", "Mean Rank",
    ]

    header_fill  = PatternFill("solid", fgColor="0D1B4B")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if Path(filepath).exists():
        wb       = load_workbook(filepath)
        ws       = wb.active
        next_row = ws.max_row + 1
        run_num  = next_row - 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Experiment Results"
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"
        next_row = 2
        run_num  = 1

    row_data = [
        run_num,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        settings.get("dataset",       "ALL"),
        settings.get("model",         "OneHot-LR"),
        settings.get("loss_func",     "LogLoss"),
        settings.get("neg_sampler",   "—"),
        settings.get("corrupt",       "—"),
        settings.get("embedding_dim", "—"),
        settings.get("num_epochs",    "—"),
        settings.get("best_epoch",    "—"),
        settings.get("batch_size",    "—"),
        settings.get("learning_rate", f"C={C}"),
        settings.get("margin",        "—"),
        settings.get("adv_temp",      "—"),
        settings.get("num_neg",       "—"),
        f"{int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s",
        eval_label,
        round(metrics.get("mrr",        0), 4),
        round(metrics.get("hits_at_1",  0), 4),
        round(metrics.get("hits_at_3",  0), 4),
        round(metrics.get("hits_at_10", 0), 4),
        round(metrics.get("mean_rank",  0), 4),
    ]

    for col, val in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col, value=val)

    wb.save(filepath)
    print(f"  Recorded: {Path(filepath).name}  Run #{run_num}")


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":

    print("=" * 60)
    print("One-Hot Encoding Logistic Regression Baseline")
    print("  Feature: concat(h_onehot, t_onehot)")
    print("=" * 60)

    # ── Load data ─────────────────────────────────────────────────────────────
    print("\nLoading triples...")
    df = pd.read_csv(CSV_PATH)
    df = df.drop_duplicates(subset=['head', 'relation', 'tail']).reset_index(drop=True)
    print(f"Clean triples : {len(df):,}")
    print(f"Relations     : {df['relation'].value_counts().to_dict()}")
    print(f"Unique heads  : {df['head'].nunique()}")
    print(f"Unique tails  : {df['tail'].nunique()}")
    print(f"Unique entities: {pd.concat([df['head'], df['tail']]).nunique()}")

    # ── Stratified split ──────────────────────────────────────────────────────
    df_train, df_val, df_test = stratified_split(
        df, train_ratio=TRAIN_RATIO, val_ratio=VAL_RATIO, random_seed=RANDOM_SEED
    )
    # Combine train + val (same as KGE scripts see during training)
    df_trainval = pd.concat([df_train, df_val]).reset_index(drop=True)

    print(f"\nTrain+Val : {len(df_trainval):,}")
    print(f"Test      : {len(df_test):,}")

    relations = sorted(df_test['relation'].unique())
    n_test    = len(df_test)

    # ── Baselines ─────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("Baselines (no model needed)")
    print("=" * 60)

    rel_counts = df_test['relation'].value_counts()
    majority   = rel_counts.idxmax()
    y_true_all = df_test['relation'].values
    y_majority = np.array([majority] * n_test)
    y_random   = np.random.RandomState(RANDOM_SEED).choice(relations, size=n_test)

    for label, y_pred in [("Majority class", y_majority), ("Random", y_random)]:
        df_per, macro = compute_per_relation_metrics(y_true_all, y_pred, relations)
        print(f"\n  {label}:")
        print(f"    {'Relation':<20} {'Count':>6} {'Hits@1':>8}")
        print(f"    {'-'*38}")
        for rel, row in df_per.iterrows():
            print(f"    {rel:<20} {row['count']:>6} {row['hits_at_1']:>8.4f}")
        print(f"    {'-'*38}")
        print(f"    {'Macro-Hits@1':<20} {'':>6} {macro:>8.4f}")

    # ── Build one-hot features ────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("Building one-hot features")
    print("=" * 60)

    # Fit encoder on ALL entities seen in the full dataset
    # (both head and tail columns) so test entities are always known
    all_entities = pd.concat([df['head'], df['tail']]).unique().reshape(-1, 1)
    n_entities   = len(all_entities)

    enc = OneHotEncoder(sparse_output=False, handle_unknown='ignore')
    enc.fit(all_entities)

    print(f"  Encoder fitted on {n_entities} unique entities")
    print(f"  Feature dim per entity : {n_entities}")
    print(f"  Feature dim per triple : {2 * n_entities}  (concat head + tail)")

    def encode_triples(df_triples):

        X_h = enc.transform(df_triples[['head']].values)  # (N, n_entities)
        X_t = enc.transform(df_triples[['tail']].values)  # (N, n_entities)
        X   = np.concatenate([X_h, X_t], axis=1)          # (N, 2*n_entities)
        y   = df_triples['relation'].values
        return X.astype(np.float32), y

    print("\n  Encoding train+val set...")
    t0 = time.time()
    X_train, y_train = encode_triples(df_trainval)
    print(f"  Encoding test set...")
    X_test,  y_test  = encode_triples(df_test)
    print(f"  Encoding done in {time.time()-t0:.2f}s")
    print(f"  X_train shape : {X_train.shape}")
    print(f"  X_test  shape : {X_test.shape}")

    # ── Train LR ──────────────────────────────────────────────────────────────
    print("=" * 60)
    print("Training Logistic Regression on one-hot features")
    print(f"  C             : {C}")
    print(f"  class_weight  : balanced")
    print(f"  max_iter      : {MAX_ITER}")
    print(f"  Features      : concat(h_onehot, t_onehot)  — {2*n_entities} dims")
    print("=" * 60)

    t0  = time.time()
    clf = LogisticRegression(
        C            = C,
        max_iter     = MAX_ITER,
        class_weight = "balanced",
        random_state = RANDOM_SEED,
        solver       = "lbfgs",
        multi_class  = "multinomial",
        n_jobs       = -1,         
    )
    clf.fit(X_train, y_train)
    elapsed = time.time() - t0
    print(f"\n  Training time : {elapsed:.2f}s")

    # ── Predict and evaluate ──────────────────────────────────────────────────
    y_pred = clf.predict(X_test)

    df_per, macro = compute_per_relation_metrics(y_test, y_pred, relations)
    micro = (y_pred == y_test).mean()

    print(f"\n{'Relation':<20} {'Count':>6} {'Correct':>8} {'Hits@1':>8}")
    print("-" * 46)
    for rel, row in df_per.iterrows():
        print(f"  {rel:<18} {row['count']:>6} {row['correct']:>8} {row['hits_at_1']:>8.4f}")
    print("-" * 46)
    print(f"  {'Macro-Hits@1':<18} {'':>6} {'':>8} {macro:>8.4f}  ← primary metric")
    print(f"  {'Micro-Hits@1':<18} {'':>6} {'':>8} {micro:>8.4f}")

    # ── Confusion matrix ──────────────────────────────────────────────────────
    cm = confusion_matrix(y_test, y_pred, labels=relations)
    print(f"\nConfusion matrix (rows=true, cols=predicted):")
    print(f"  Labels: {relations}")
    col_w = max(len(r) for r in relations) + 2
    print(f"  {'True || Pred':<20}", end="")
    for r in relations:
        print(f"  {r:>{col_w}}", end="")
    print()
    print(f"  {'-' * (20 + (col_w + 2) * len(relations))}")
    for i, row in enumerate(cm):
        print(f"  {relations[i]:<20}", end="")
        for val in row:
            print(f"  {val:>{col_w}}", end="")
        print()

    # ── Save per-triple results ───────────────────────────────────────────────
    df_eval = df_test.copy().reset_index(drop=True)
    df_eval['predicted'] = y_pred
    df_eval['true_label']= y_test
    df_eval['correct']   = (y_pred == y_test)

    eval_path = OUTPUT_DIR / "onehot_lr_eval.csv"
    df_eval[['head', 'relation', 'tail', 'domain',
             'true_label', 'predicted', 'correct']].to_csv(eval_path, index=False)
    print(f"\nSaved per-triple results: {eval_path.name}  ({len(df_eval):,} triples)")

    # Quick breakdown of mistakes
    wrong    = df_eval[~df_eval['correct']]
    wrong_ct = df_eval[
        (df_eval['true_label'] == 'CONTRARY_TO') & (~df_eval['correct'])
    ]
    print(f"\nTotal mistakes     : {len(wrong):,} / {len(df_eval):,}")
    print(f"CONTRARY_TO missed : {len(wrong_ct):,} / "
          f"{(df_eval['true_label']=='CONTRARY_TO').sum():,}")
    if len(wrong_ct) > 0:
        print(f"  Predicted instead:")
        print(wrong_ct['predicted'].value_counts().to_string())

    # ── Record to Excel ───────────────────────────────────────────────────────
    settings = {
        "dataset":       "ALL",
        "model":         "OneHot-LR",
        "loss_func":     "LogLoss",
        "embedding_dim": f"OneHot ({2*n_entities})",
        "learning_rate": f"C={C}",
    }
    metrics_out = {
        "mrr":        macro,   # macro-H@1 stored as primary metric
        "hits_at_1":  micro,
        "hits_at_3":  1.0,     # trivially 1.0 with 3 relations
        "hits_at_10": 1.0,
        "mean_rank":  0.0,     # not applicable
    }
    record_results(RESULTS_FILE, settings, metrics_out, elapsed,
                   eval_label="OneHot-LR")

    # ── One-Hot LR vs KGE comparison ─────────────────────────────────────────
    print("\n" + "=" * 60)
    print("One-Hot LR vs KGE Comparison  (Macro-Hits@1)")
    print("=" * 60)

    try:
        df_xl = pd.read_excel(RESULTS_FILE)

        # KGE rows — Evaluator ends with "-Macro" (written by KGE scripts)
        kge_rows = df_xl[df_xl["Evaluator"].str.endswith("-Macro", na=False)][
            ["Model", "MRR"]
        ].copy()
        kge_rows.columns = ["Model", "Macro-H@1"]
        kge_rows = kge_rows.drop_duplicates("Model").sort_values("Macro-H@1", ascending=False)

        # LR-KGE rows (from lr_baseline.py runs, any source model)
        lr_kge_rows = df_xl[df_xl["Evaluator"].str.startswith("LR-", na=False)][
            ["Model", "MRR"]
        ].copy()
        lr_kge_rows.columns = ["Model", "Macro-H@1"]
        lr_kge_rows = (lr_kge_rows.groupby("Model")["Macro-H@1"]
                       .max().reset_index()
                       .sort_values("Macro-H@1", ascending=False))

        # One-Hot LR result from this run
        onehot_rows = df_xl[df_xl["Evaluator"] == "OneHot-LR"][["Model", "MRR"]].copy()
        onehot_rows.columns = ["Model", "Macro-H@1"]

        best_onehot = macro   # from current run

        print(f"\n  {'Model':<30} {'Macro-H@1':>10}  vs One-Hot")
        print(f"  {'-'*52}")

        # Print KGE models
        if len(kge_rows):
            print(f"  --- KGE models ---")
            for _, row in kge_rows.iterrows():
                diff   = row["Macro-H@1"] - best_onehot
                marker = f"  {'better' if diff > 0.001 else ('≈ tied' if abs(diff) <= 0.001 else 'worse')}  ({diff:+.4f})"
                print(f"  {row['Model']:<30} {row['Macro-H@1']:>10.4f}{marker}")

        # Print LR-on-KGE best per source model
        if len(lr_kge_rows):
            print(f"  --- LR on KGE embeddings ---")
            for _, row in lr_kge_rows.iterrows():
                diff   = row["Macro-H@1"] - best_onehot
                marker = f"  {'better' if diff > 0.001 else ('≈ tied' if abs(diff) <= 0.001 else 'worse')}  ({diff:+.4f})"
                print(f"  {row['Model']:<30} {row['Macro-H@1']:>10.4f}{marker}")

        # Print One-Hot result
        print(f"  --- One-Hot LR ---")
        print(f"  {'OneHot-LR':<30} {best_onehot:>10.4f}")
        print(f"  {'Majority baseline':<30} {'0.3333':>10}")

        # Summary verdict
        best_kge_macro = kge_rows["Macro-H@1"].max() if len(kge_rows) else None
        if best_kge_macro is not None:
            gap = best_kge_macro - best_onehot
            print(f"\n  Best KGE : {best_kge_macro:.4f}  |  One-Hot LR : {best_onehot:.4f}  |  Gap : {gap:+.4f}")


    except Exception as e:
        print(f"  [Could not load comparison from Excel: {e}]")
        print(f"  One-Hot LR Macro-Hits@1 = {macro:.4f}")

    print("\n" + "=" * 60)
    print("All done.")
    print("=" * 60)