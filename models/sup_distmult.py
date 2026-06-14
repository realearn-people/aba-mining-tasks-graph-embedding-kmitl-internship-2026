"""
For 3-relation training (CONTRARY_TO, NOT_CONTRARY, SUPPORT)

"""

import pandas as pd
import numpy as np
import torch
import time
import datetime
import pickle
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pykeen.pipeline import pipeline
from pykeen.triples import TriplesFactory
from pykeen.sampling import BasicNegativeSampler
from sklearn.decomposition import PCA
from sklearn.model_selection import train_test_split

# ── File paths ────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
CSV_PATH     = ROOT / "data" / "input_data" / "hotel_contrary_dataset_support.csv"   # input dataset
OUTPUT_DIR   = ROOT / "outputs" / "distmult_sup_output"             # folder for saved outputs
RESULTS_FILE = ROOT / "data" / "experiment_results" / "sup_experiment_results.xlsx"           # auto-recorded experiment log
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Hyperparameters ───────────────────────────────────────────────────────────

DATASET            = "ALL"
EMBEDDING_DIM      = 100
NUM_EPOCHS         = 1000
BATCH_SIZE         = 256
LEARNING_RATE      = 0.0001
MARGIN             = 1.0
TRAIN_RATIO        = 0.8
VAL_RATIO          = 0.1
RANDOM_SEED        = 42
NUM_NEG            = 3
MODEL              = "DistMult"
LOSS               = "NSSALoss"
ADV_TEMP           = 0.5
CORRUPT            = "relation"
REGULARIZER_WEIGHT = 1e-3       

# ── ABA claim roots ───────────────────────────────────────────────────────────
CLAIMS = [
    'good_staff',    'bad_staff',
    'good_price',    'bad_price',
    'good_check-in', 'bad_check-in',
    'good_check-out','bad_check-out',
]


# ── Extended negative sampler ─────────────────────────────────────────────────
class ExtendedBasicNegativeSampler(BasicNegativeSampler):

    def __init__(self, *args, save_path=str(OUTPUT_DIR / "negative_samples.csv"), **kwargs):
        kwargs['corruption_scheme'] = ['relation']
        super().__init__(*args, **kwargs)
        object.__setattr__(self, 'save_path', save_path)
        object.__setattr__(self, 'header_written', False)
        object.__setattr__(self, 'batch_count', 0)

    def corrupt_batch(self, positive_batch: torch.LongTensor) -> torch.LongTensor:
        negatives = super().corrupt_batch(positive_batch=positive_batch)
        object.__setattr__(self, 'batch_count', self.batch_count + 1)

        if self.batch_count == 1:
            print(f"[DEBUG] corrupt_batch called for the first time!")
            print(f"[DEBUG] Negative shape: {negatives.shape}")
        if self.batch_count % 100 == 0:
            print(f"[DEBUG] Batches captured so far: {self.batch_count}")

        try:
            neg_flat = negatives.detach().cpu().reshape(-1, 3)
            rows = []
            for t in neg_flat:
                try:
                    rows.append({
                        "head":     id_to_entity.get(t[0].item(), "unknown"),
                        "relation": id_to_relation.get(t[1].item(), "unknown"),
                        "tail":     id_to_entity.get(t[2].item(), "unknown"),
                        "label":    0,
                    })
                except Exception as row_err:
                    print(f"[WARN] Skipping row in batch {self.batch_count}: {row_err}")
                    continue
            if rows:
                df_batch = pd.DataFrame(rows)
                df_batch.to_csv(
                    self.save_path,
                    mode="w" if not self.header_written else "a",
                    header=not self.header_written,
                    index=False,
                )
                object.__setattr__(self, 'header_written', True)
                del df_batch
            del neg_flat, rows
        except Exception as batch_err:
            print(f"[WARN] Skipping batch {self.batch_count}: {batch_err}")
        return negatives


# ══════════════════════════════════════════════════════════════════════════════
# Relation-rank evaluator
# ══════════════════════════════════════════════════════════════════════════════
class RelationRankEvaluator:


    def __init__(self, model, triples_factory, batch_size=256, device=None):
        self.model           = model
        self.triples_factory = triples_factory
        self.num_relations   = triples_factory.num_relations
        self.batch_size      = batch_size
        self.device          = device or ("cuda" if torch.cuda.is_available() else "cpu")

    def evaluate(self, mapped_triples, additional_filter_triples=None):
        self.model.eval()
        self.model.to(self.device)

        filter_map = {}
        if additional_filter_triples:
            for triples in additional_filter_triples:
                for triple in triples:
                    h, r, t = triple[0].item(), triple[1].item(), triple[2].item()
                    filter_map.setdefault((h, t), set()).add(r)

        real_rel_ids = torch.tensor(
            [rid for label, rid in self.triples_factory.relation_to_id.items()
             if not label.endswith("_inverse")],
            dtype=torch.long,
            device=self.device,
        )
        num_real_rels = len(real_rel_ids)
        print(f"  Relations in vocab : {self.num_relations} "
              f"({num_real_rels} real — ranking among these)")

        eval_rows = []
        print(f"  Evaluating {len(mapped_triples):,} test triples...")

        with torch.no_grad():
            for i in range(0, len(mapped_triples), self.batch_size):
                batch = mapped_triples[i : i + self.batch_size]

                for triple in batch:
                    h_id = triple[0].item()
                    r_id = triple[1].item()
                    t_id = triple[2].item()

                    h_tensor = torch.full((num_real_rels,), h_id, device=self.device)
                    t_tensor = torch.full((num_real_rels,), t_id, device=self.device)
                    hrt_all  = torch.stack([h_tensor, real_rel_ids, t_tensor], dim=1)
                    scores   = self.model.score_hrt(hrt_all).squeeze()

                    masked_scores = scores.clone()
                    known_rels    = filter_map.get((h_id, t_id), set())
                    for local_idx, global_rid in enumerate(real_rel_ids.tolist()):
                        if global_rid != r_id and global_rid in known_rels:
                            masked_scores[local_idx] = float('-inf')

                    true_local_idx = (real_rel_ids == r_id).nonzero(as_tuple=True)[0].item()
                    true_score     = masked_scores[true_local_idx].item()
                    rank           = int((masked_scores > true_score).sum().item()) + 1

                    eval_rows.append({
                        "head":            id_to_entity.get(h_id, "unknown"),
                        "relation":        id_to_relation.get(r_id, "unknown"),
                        "tail":            id_to_entity.get(t_id, "unknown"),
                        "relation_rank":   rank,
                        "reciprocal_rank": round(1.0 / rank, 6),
                        "hits_at_1":       int(rank <= 1),
                        "hits_at_3":       int(rank <= 3),
                        "hits_at_10":      int(rank <= 10),
                    })

                if (i // self.batch_size + 1) % 20 == 0:
                    done = min(i + self.batch_size, len(mapped_triples))
                    print(f"  Processed {done:,} / {len(mapped_triples):,}")

        # ── Aggregate metrics ──────────────────────────────────────────────────
        df_eval = pd.DataFrame(eval_rows)
        mrr     = df_eval["reciprocal_rank"].mean()
        h_at_1  = df_eval["hits_at_1"].mean()
        h_at_3  = df_eval["hits_at_3"].mean()
        h_at_10 = df_eval["hits_at_10"].mean()
        mr      = df_eval["relation_rank"].mean()

        # ── Per-relation breakdown ─────────────────────────────────────────────
        per_rel = (
            df_eval.groupby("relation")
            .agg(
                count     =("hits_at_1", "count"),
                mrr       =("reciprocal_rank", "mean"),
                hits_at_1 =("hits_at_1", "mean"),
                mean_rank =("relation_rank", "mean"),
            )
            .round(4)
        )
        macro_h1 = per_rel["hits_at_1"].mean()

        print(f"\n{'Relation':<20} {'Count':>6} {'MRR':>7} {'Hits@1':>8} {'MeanRank':>10}")
        print("-" * 56)
        for rel, row in per_rel.iterrows():
            print(f"  {rel:<18} {int(row['count']):>6} {row['mrr']:>7.4f} "
                  f"{row['hits_at_1']:>8.4f} {row['mean_rank']:>10.4f}")
        print("-" * 56)
        print(f"  {'Macro-avg':<18} {'':>6} {'':>7} {macro_h1:>8.4f}")
        print(f"\n  Hits@3={h_at_3:.4f} / Hits@10={h_at_10:.4f} trivially 1.0 "
              f"with {len(per_rel)} relations — ignore them.")
        print(f"  Primary metrics: per-relation Hits@1 and Macro-Hits@1.\n")

        metrics_dict = {
            "both": {
                "realistic": {
                    "inverse_harmonic_mean_rank": round(mrr,      4),
                    "arithmetic_mean_rank":       round(mr,       4),
                    "hits_at_1":                  round(h_at_1,   4),
                    "hits_at_3":                  round(h_at_3,   4),
                    "hits_at_10":                 round(h_at_10,  4),
                    "macro_hits_at_1":            round(macro_h1, 4),
                }
            },
            "per_relation": per_rel.to_dict(orient="index"),
        }

        return metrics_dict, df_eval


# ── Experiment recorder ───────────────────────────────────────────────────────
def record_results(filepath, settings, metrics, elapsed, eval_label="PyKEEN"):
    headers = [
        "No.", "Timestamp", "Dataset", "Model", "Loss Function",
        "Neg Sampler", "Corruption", "Embedding Dim", "Num Epochs",
        "Best Epoch", "Batch Size", "Learning Rate", "Margin",
        "Adv Temp", "Num Neg", "Training Time", "Evaluator",
        "MRR", "Hits@1", "Hits@3", "Hits@10", "Mean Rank",
    ]

    header_fill  = PatternFill("solid", fgColor="0D1B4B")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if Path(filepath).exists():
        wb       = load_workbook(filepath)
        ws       = wb.active
        next_row = ws.max_row + 1
        run_num  = next_row - 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Experiment Results"
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"
        next_row = 2
        run_num  = 1

    row_data = [
        run_num,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        settings["dataset"],
        settings["model"],
        settings["loss_func"],
        settings["neg_sampler"],
        settings["corrupt"],
        settings["embedding_dim"],
        settings["num_epochs"],
        settings.get("best_epoch", "—"),
        settings["batch_size"],
        settings["learning_rate"],
        settings["margin"],
        settings["adv_temp"],
        settings["num_neg"],
        f"{int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s",
        eval_label,
        round(metrics["mrr"],        4),
        round(metrics["hits_at_1"],  4),
        round(metrics["hits_at_3"],  4),
        round(metrics["hits_at_10"], 4),
        round(metrics["mean_rank"],  4),
    ]

    for col, val in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col, value=val)

    wb.save(filepath)
    print(f"Recorded: {filepath}  Run #{run_num}")


# ── Stratified split ──────────────────────────────────────────────────────────
def stratified_split(df, train_ratio=0.8, val_ratio=0.1, random_seed=42):
    train_dfs, val_dfs, test_dfs = [], [], []

    for (domain, relation), group in df.groupby(['domain', 'relation']):
        n = len(group)
        if n < 10:
            print(f"  [WARN] {domain} | {relation} — only {n} rows, added to train")
            train_dfs.append(group)
            continue

        train_val, test = train_test_split(
            group,
            test_size=1 - train_ratio - val_ratio,
            random_state=random_seed,
        )
        val_size = val_ratio / (train_ratio + val_ratio)
        train, val = train_test_split(
            train_val,
            test_size=val_size,
            random_state=random_seed,
        )

        train_dfs.append(train)
        val_dfs.append(val)
        test_dfs.append(test)

        print(f"  {domain:12} | {relation:15} — "
              f"train: {len(train):,}  val: {len(val):,}  test: {len(test):,}")

    df_train = pd.concat(train_dfs).sample(frac=1, random_state=random_seed).reset_index(drop=True)
    df_val   = pd.concat(val_dfs  ).sample(frac=1, random_state=random_seed).reset_index(drop=True)
    df_test  = pd.concat(test_dfs ).sample(frac=1, random_state=random_seed).reset_index(drop=True)

    return df_train, df_val, df_test


# ── ABA node level classifier ─────────────────────────────────────────────────
def get_node_level(entity_name):
    if entity_name in CLAIMS:
        return 'claim'
    elif entity_name.startswith('no_evident_'):
        return 'attacker'
    else:
        return 'body'


# ── ABA tree builder ──────────────────────────────────────────────────────────
def build_aba_tree(df_orig, claims):
    tree_rows = []
    for claim in claims:
        bodies = df_orig[
            (df_orig['tail'] == claim) &
            (df_orig['relation'] == 'SUPPORT')
        ]
        for _, b_row in bodies.iterrows():
            body       = b_row['head']
            score_body = b_row['score']
            attackers  = df_orig[
                (df_orig['tail'] == body) &
                (df_orig['relation'].isin(['CONTRARY_TO', 'NOT_CONTRARY']))
            ]
            if len(attackers) == 0:
                tree_rows.append({
                    'claim':           claim,
                    'body':            body,
                    'attacker':        None,
                    'body_relation':   'SUPPORT',
                    'attack_relation': None,
                    'body_score':      round(score_body, 6),
                    'attack_score':    None,
                    'domain':          b_row.get('domain', ''),
                    'body_level':      get_node_level(body),
                })
            else:
                for _, a_row in attackers.iterrows():
                    tree_rows.append({
                        'claim':           claim,
                        'body':            body,
                        'attacker':        a_row['head'],
                        'body_relation':   'SUPPORT',
                        'attack_relation': a_row['relation'],
                        'body_score':      round(score_body, 6),
                        'attack_score':    round(a_row['score'], 6),
                        'domain':          a_row.get('domain', ''),
                        'body_level':      get_node_level(body),
                    })
    return pd.DataFrame(tree_rows)


# ══════════════════════════════════════════════════════════════════════════════
# Main training pipeline
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":

    # ── Load data ─────────────────────────────────────────────────────────────
    print("=" * 60)
    print("Loading triples")
    print("=" * 60)

    df = pd.read_csv(CSV_PATH)
    df_raw = df.copy()
    df = df.drop_duplicates(subset=['head', 'relation', 'tail']).reset_index(drop=True)
    print(f"Raw rows   : {len(df_raw):,}")
    print(f"Clean rows : {len(df):,}")
    print(f"Dropped    : {len(df_raw) - len(df):,}")
    print(f"Relations found: {df['relation'].value_counts().to_dict()}")

    # ── Stratified split ──────────────────────────────────────────────────────
    print("=" * 60)
    print("Stratified split by domain + relation")
    print("=" * 60)

    df_train, df_val, df_test = stratified_split(
        df,
        train_ratio=TRAIN_RATIO,
        val_ratio=VAL_RATIO,
        random_seed=RANDOM_SEED,
    )

    print(f"\nTrain : {len(df_train):,}")
    print(f"Val   : {len(df_val):,}")
    print(f"Test  : {len(df_test):,}")

    print("\nDomain x Relation distribution in test set:")
    print(df_test.groupby(['domain', 'relation']).size().to_string())

    training = TriplesFactory.from_labeled_triples(
        triples=df_train[['head', 'relation', 'tail']].values.astype(str),
    )
    validation = TriplesFactory.from_labeled_triples(
        triples=df_val[['head', 'relation', 'tail']].values.astype(str),
        entity_to_id=training.entity_to_id,
        relation_to_id=training.relation_to_id,
    )
    testing = TriplesFactory.from_labeled_triples(
        triples=df_test[['head', 'relation', 'tail']].values.astype(str),
        entity_to_id=training.entity_to_id,
        relation_to_id=training.relation_to_id,
    )

    print(f"\nUnique entities  : {training.num_entities:,}")
    print(f"Unique relations : {training.num_relations}")

    id_to_entity   = {v: k for k, v in training.entity_to_id.items()}
    id_to_relation = {v: k for k, v in training.relation_to_id.items()}

    # ── Train DistMult ────────────────────────────────────────────────────────
    print("=" * 60)
    print(f"Training {MODEL}")
    print("=" * 60)

    start_time = time.time()

    result = pipeline(
        training=training,
        validation=validation,
        testing=testing,

        # ── Model ─────────────────────────────────────────────────────────────
        model=MODEL,
        model_kwargs=dict(
            embedding_dim      = EMBEDDING_DIM,
            regularizer        = "LpRegularizer",
            regularizer_kwargs = dict(
                weight    = REGULARIZER_WEIGHT,
                p         = 2.0,
                normalize = False,
            ),
        ),

        # ── Loss ──────────────────────────────────────────────────────────────
        loss=LOSS,
        loss_kwargs=dict(
            margin=MARGIN,
            adversarial_temperature=ADV_TEMP,
        ),

        # ── Negative sampler: relation corruption ─────────────────────────────
        negative_sampler=ExtendedBasicNegativeSampler,
        negative_sampler_kwargs=dict(
            num_negs_per_pos=NUM_NEG,
            corruption_scheme=[CORRUPT],
        ),

        # ── Optimizer ─────────────────────────────────────────────────────────
        optimizer="Adam",
        optimizer_kwargs=dict(lr=LEARNING_RATE),

        # ── Training loop ─────────────────────────────────────────────────────
        training_loop="sLCWA",
        training_kwargs=dict(
            num_epochs=NUM_EPOCHS,
            batch_size=BATCH_SIZE,
        ),

        # ── Early stopping ─────────────────────────────────────────────────────
        stopper="early",
        stopper_kwargs=dict(
            metric="hits_at_10",
            patience=10,
            relative_delta=0.002,
            frequency=10,
        ),

        # ── Evaluator ─────────────────────────────────────────────────────────
        evaluator="RankBasedEvaluator",
        evaluator_kwargs=dict(filtered=True),

        random_seed=RANDOM_SEED,
        device="cuda" if torch.cuda.is_available() else "cpu",
    )

    end_time = time.time()
    elapsed  = end_time - start_time
    print(f"\nTraining completed in: "
          f"{int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s")

    # ── PyKEEN default evaluation ─────────────────────────────────────────────
    print("=" * 60)
    print("Evaluation Results (PyKEEN default — entity ranking)")
    print("=" * 60)

    metrics_dict = result.metric_results.to_dict()
    r_default    = metrics_dict["both"]["realistic"]

    mrr        = r_default['inverse_harmonic_mean_rank']
    hits_at_1  = r_default['hits_at_1']
    hits_at_3  = r_default['hits_at_3']
    hits_at_10 = r_default['hits_at_10']
    mean_rank  = r_default['arithmetic_mean_rank']

    print(f"MRR       : {mrr:.4f}")
    print(f"Hits@1    : {hits_at_1:.4f}")
    print(f"Hits@3    : {hits_at_3:.4f}")
    print(f"Hits@10   : {hits_at_10:.4f}")
    print(f"Mean Rank : {mean_rank:.4f}")

    # ── Record PyKEEN results ─────────────────────────────────────────────────
    try:
        best_epoch = result.stopper.best_epoch
    except Exception:
        best_epoch = "—"

    settings = {
        "dataset":       DATASET,
        "model":         MODEL,
        "loss_func":     LOSS,
        "neg_sampler":   "ExtendedBasicNegativeSampler",
        "corrupt":       CORRUPT,
        "embedding_dim": EMBEDDING_DIM,
        "num_epochs":    NUM_EPOCHS,
        "best_epoch":    best_epoch,
        "batch_size":    BATCH_SIZE,
        "learning_rate": LEARNING_RATE,
        "margin":        MARGIN,
        "adv_temp":      ADV_TEMP,
        "num_neg":       NUM_NEG,
    }

    metrics_out = {
        "mrr":        mrr,
        "hits_at_1":  hits_at_1,
        "hits_at_3":  hits_at_3,
        "hits_at_10": hits_at_10,
        "mean_rank":  mean_rank,
    }

    record_results(RESULTS_FILE, settings, metrics_out, elapsed, eval_label="PyKEEN")

    # ── Relation-rank evaluation ──────────────────────────────────────────────
    print("=" * 60)
    print("Relation-Rank Evaluation")
    print("=" * 60)

    try:
        rel_evaluator = RelationRankEvaluator(
            model=result.model,
            triples_factory=training,
            batch_size=BATCH_SIZE,
            device="cuda" if torch.cuda.is_available() else "cpu",
        )

        rel_metrics_dict, df_rel_eval = rel_evaluator.evaluate(
            mapped_triples=testing.mapped_triples,
            additional_filter_triples=[
                training.mapped_triples,
                validation.mapped_triples,
            ],
        )

        r_rel     = rel_metrics_dict["both"]["realistic"]
        rel_mrr   = r_rel["inverse_harmonic_mean_rank"]
        rel_h1    = r_rel["hits_at_1"]
        rel_h3    = r_rel["hits_at_3"]
        rel_h10   = r_rel["hits_at_10"]
        rel_mr    = r_rel["arithmetic_mean_rank"]
        rel_macro = r_rel.get("macro_hits_at_1", float("nan"))

        print(f"\n{'Metric':<16} {'RelationRank':>14}  {'PyKEEN (entity)':>16}")
        print("-" * 50)
        print(f"{'MRR':<16} {rel_mrr:>14.4f}  {mrr:>16.4f}")
        print(f"{'Hits@1 (micro)':<16} {rel_h1:>14.4f}  {hits_at_1:>16.4f}")
        print(f"{'Hits@1 (macro)':<16} {rel_macro:>14.4f}  {'N/A':>16}")
        print(f"{'Hits@3':<16} {rel_h3:>14.4f}  {hits_at_3:>16.4f}  <- trivially 1.0")
        print(f"{'Hits@10':<16} {rel_h10:>14.4f}  {hits_at_10:>16.4f}  <- trivially 1.0")
        print(f"{'Mean Rank':<16} {rel_mr:>14.4f}  {mean_rank:>16.4f}")

        metrics_rel_out = {
            "mrr":        rel_mrr,
            "hits_at_1":  rel_h1,
            "hits_at_3":  rel_h3,
            "hits_at_10": rel_h10,
            "mean_rank":  rel_mr,
        }
        record_results(
            RESULTS_FILE, settings, metrics_rel_out, elapsed,
            eval_label="RelationRank"
        )
        record_results(
            RESULTS_FILE, settings,
            {"mrr": rel_macro, "hits_at_1": rel_macro,
             "hits_at_3": rel_macro, "hits_at_10": rel_macro,
             "mean_rank": rel_mr},
            elapsed, eval_label="RelationRank-Macro"
        )

        # Join domain from original CSV
        df_orig_for_join = pd.read_csv(CSV_PATH)[
            ['head', 'relation', 'tail', 'domain']
        ].drop_duplicates(subset=['head', 'relation', 'tail'])
        df_rel_eval = df_rel_eval.merge(
            df_orig_for_join,
            on=['head', 'relation', 'tail'],
            how='left'
        )

        rel_eval_path = OUTPUT_DIR / "relation_eval_ranked.csv"
        df_rel_eval.to_csv(rel_eval_path, index=False, mode='w')
        print(f"\nSaved: {rel_eval_path}  ({len(df_rel_eval):,} triples)")

    except Exception as e:
        print(f"[ERROR] Relation-rank evaluation failed: {e}")
        import traceback; traceback.print_exc()

    # ── Export visualization CSV + ABA tree structure ─────────────────────────
    print("=" * 60)
    print("Exporting visualization data + ABA tree structure")
    print("=" * 60)

    try:
        # DistMult embeddings are real-valued — straightforward extraction,
        # no complex parts to handle (unlike RotatE and ComplEx).
        entity_emb = result.model.entity_representations[0](
            indices=None
        ).detach().cpu().numpy().astype(np.float32)

        print(f"Entity embedding shape: {entity_emb.shape}")

        pca       = PCA(n_components=2, random_state=RANDOM_SEED)
        coords_2d = pca.fit_transform(entity_emb)
        print(f"PCA explained variance ratio: {pca.explained_variance_ratio_.round(4)}")

        entity_rows = []
        for entity_name, entity_id in training.entity_to_id.items():
            row_data = {
                "entity_id":   entity_id,
                "entity_name": entity_name,
                "x":           round(float(coords_2d[entity_id, 0]), 6),
                "y":           round(float(coords_2d[entity_id, 1]), 6),
            }
            for dim_idx, val in enumerate(entity_emb[entity_id]):
                row_data[f"emb_{dim_idx}"] = round(float(val), 6)
    
            entity_rows.append(row_data) 
        df_entities = pd.DataFrame(entity_rows)

        # Score all original triples
        df_orig     = pd.read_csv(CSV_PATH)
        scores_list = []
        result.model.eval()
        with torch.no_grad():
            for _, row in df_orig.iterrows():
                try:
                    h_id = training.entity_to_id.get(str(row['head']), -1)
                    r_id = training.relation_to_id.get(str(row['relation']), -1)
                    t_id = training.entity_to_id.get(str(row['tail']), -1)
                    if -1 in (h_id, r_id, t_id):
                        scores_list.append(None)
                        continue
                    hrt = torch.tensor([[h_id, r_id, t_id]], dtype=torch.long)
                    s   = result.model.score_hrt(hrt)
                    scores_list.append(round(float(s.item()), 6))
                except Exception:
                    scores_list.append(None)

        df_orig['score'] = scores_list

        coord_map         = df_entities.set_index('entity_name')[['x', 'y']]
        df_orig['head_x'] = df_orig['head'].map(coord_map['x'])
        df_orig['head_y'] = df_orig['head'].map(coord_map['y'])
        df_orig['tail_x'] = df_orig['tail'].map(coord_map['x'])
        df_orig['tail_y'] = df_orig['tail'].map(coord_map['y'])
        df_orig = df_orig.dropna(
            subset=['score', 'head_x', 'head_y', 'tail_x', 'tail_y']
        )

        VIZ_FILE = OUTPUT_DIR / "visualization_data.csv"
        df_orig[['head', 'relation', 'tail', 'score',
                 'head_x', 'head_y', 'tail_x', 'tail_y']].to_csv(
            VIZ_FILE, index=False
        )
        print(f"Saved: {VIZ_FILE}  ({len(df_orig):,} triples)")

        df_tree   = build_aba_tree(df_orig, CLAIMS)
        TREE_FILE = OUTPUT_DIR / "aba_tree_structure.csv"
        df_tree.to_csv(TREE_FILE, index=False)
        print(f"Saved: {TREE_FILE}  ({len(df_tree):,} rows)")
        print(f"  Claims:    {df_tree['claim'].nunique()}")
        print(f"  Bodies:    {df_tree['body'].nunique()}")
        print(f"  Attackers: {df_tree['attacker'].nunique()}")

        ENTITY_FILE = OUTPUT_DIR / "entity_embeddings.csv"
        df_entities.to_csv(ENTITY_FILE, index=False)
        print(f"Saved: {ENTITY_FILE}  ({len(df_entities):,} entities)")

    except Exception as e:
        print(f"[ERROR] Visualization export failed: {e}")
        import traceback; traceback.print_exc()

    model_save_path = OUTPUT_DIR / f"trained_model_{MODEL}.pkl"
    tf_save_path    = OUTPUT_DIR / f"triples_factory_{MODEL}.pkl"
    
    # Save trained model weights
    with open(model_save_path, "wb") as f:
        pickle.dump(result.model, f)
    print(f"Saved model to {model_save_path}")
    
    # Save TriplesFactory (entity/relation ID mappings)
    with open(tf_save_path, "wb") as f:
        pickle.dump(training, f)
    print(f"Saved TF to {tf_save_path}")
    df_test[['head','relation','tail']].to_csv(OUTPUT_DIR / "test_triples.csv", index=False)
    
    print("=" * 60)
    print("All done.")
    print("=" * 60)