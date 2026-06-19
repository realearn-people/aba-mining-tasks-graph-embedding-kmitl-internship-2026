"""
Logistic Regression baseline for ABA contrary-mining relation prediction.

Claude-Assisted
"""

import pandas as pd
import numpy as np
import datetime
import time
import warnings
from pathlib import Path

from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.model_selection import train_test_split
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings("ignore", category=UserWarning)

# ── Paths ─────────────────────────────────────────────────────────────────────
ROOT         = Path(__file__).resolve().parent.parent
RESULTS_FILE = ROOT / "data" / "experiment_results" / "sup_experiment_results.xlsx"
CSV_PATH     = ROOT / "data" / "input_data" / "hotel_contrary_dataset_support.csv"

# Change this to: "rotate", "transe", "convkb", "complex", "distmult"
SOURCE_MODEL = "distmult"
EMB_PATH     = ROOT / "outputs" / f"{SOURCE_MODEL}_sup_output" / "entity_embeddings.csv"

# ── Hyperparameters ───────────────────────────────────────────────────────────
TRAIN_RATIO  = 0.8
VAL_RATIO    = 0.1
RANDOM_SEED  = 42
MAX_ITER     = 1000       # solver iterations 
C            = 1.0        # inverse regularization strength 

# Feature strategies to evaluate 
FEATURE_STRATEGIES = ["concat", "diff", "hadamard", "all"]

# ── ABA claims ────────────────────────────────────────────────────────────────
CLAIMS = [
    'good_staff',    'bad_staff',
    'good_price',    'bad_price',
    'good_check-in', 'bad_check-in',
    'good_check-out','bad_check-out',
]


# ── Stratified split (identical to KGE scripts) ───────────────────────────────
def stratified_split(df, train_ratio=0.8, val_ratio=0.1, random_seed=42):
    train_dfs, val_dfs, test_dfs = [], [], []

    for (domain, relation), group in df.groupby(['domain', 'relation']):
        n = len(group)
        if n < 10:
            train_dfs.append(group)
            continue

        train_val, test = train_test_split(
            group,
            test_size=1 - train_ratio - val_ratio,
            random_state=random_seed,
        )
        val_size = val_ratio / (train_ratio + val_ratio)
        train, val = train_test_split(
            train_val,
            test_size=val_size,
            random_state=random_seed,
        )

        train_dfs.append(train)
        val_dfs.append(val)
        test_dfs.append(test)

    df_train = pd.concat(train_dfs).sample(frac=1, random_state=random_seed).reset_index(drop=True)
    df_val   = pd.concat(val_dfs  ).sample(frac=1, random_state=random_seed).reset_index(drop=True)
    df_test  = pd.concat(test_dfs ).sample(frac=1, random_state=random_seed).reset_index(drop=True)

    return df_train, df_val, df_test


# ── Feature builder ───────────────────────────────────────────────────────────
def build_features(df_triples, emb_lookup, strategy="concat"):
    X_rows, y_rows, valid_idx = [], [], []

    for idx, row in df_triples.iterrows():
        h_name = str(row['head'])
        t_name = str(row['tail'])

        if h_name not in emb_lookup or t_name not in emb_lookup:
            continue  # entity not in training vocab — skip

        h = emb_lookup[h_name]
        t = emb_lookup[t_name]

        if strategy == "concat":
            feat = np.concatenate([h, t])
        elif strategy == "diff":
            feat = h - t
        elif strategy == "hadamard":
            feat = h * t
        elif strategy == "all":
            feat = np.concatenate([h, t, h - t, h * t])
        else:
            raise ValueError(f"Unknown strategy: {strategy}")

        X_rows.append(feat)
        y_rows.append(str(row['relation']))
        valid_idx.append(idx)

    X = np.array(X_rows, dtype=np.float32)
    y = np.array(y_rows)
    return X, y, valid_idx


# ── Per-relation metrics ──────────────────────────────────────────────────────
def compute_per_relation_metrics(y_true, y_pred, relations=None):
    if relations is None:
        relations = sorted(set(y_true))

    rows = []
    for rel in relations:
        mask      = y_true == rel
        count     = mask.sum()
        correct   = ((y_true == rel) & (y_pred == rel)).sum()
        hits_at_1 = correct / count if count > 0 else 0.0
        rows.append({
            "relation": rel,
            "count":    int(count),
            "correct":  int(correct),
            "hits_at_1": round(hits_at_1, 4),
        })

    df_per = pd.DataFrame(rows).set_index("relation")
    macro  = df_per["hits_at_1"].mean()
    return df_per, round(macro, 4)


# ── Experiment recorder ───────────────────────────────────────────────────────
def record_results(filepath, settings, metrics, elapsed, eval_label="LR"):
    headers = [
        "No.", "Timestamp", "Dataset", "Model", "Loss Function",
        "Neg Sampler", "Corruption", "Embedding Dim", "Num Epochs",
        "Best Epoch", "Batch Size", "Learning Rate", "Margin",
        "Adv Temp", "Num Neg", "Training Time", "Evaluator",
        "MRR", "Hits@1", "Hits@3", "Hits@10", "Mean Rank",
    ]

    header_fill  = PatternFill("solid", fgColor="0D1B4B")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if Path(filepath).exists():
        wb       = load_workbook(filepath)
        ws       = wb.active
        next_row = ws.max_row + 1
        run_num  = next_row - 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Experiment Results"
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"
        next_row = 2
        run_num  = 1

    row_data = [
        run_num,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        settings.get("dataset", "ALL"),
        settings.get("model",   "—"),
        settings.get("loss_func",    "—"),
        settings.get("neg_sampler",  "—"),
        settings.get("corrupt",      "—"),
        settings.get("embedding_dim","—"),
        settings.get("num_epochs",   "—"),
        settings.get("best_epoch",   "—"),
        settings.get("batch_size",   "—"),
        settings.get("learning_rate","—"),
        settings.get("margin",       "—"),
        settings.get("adv_temp",     "—"),
        settings.get("num_neg",      "—"),
        f"{int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s",
        eval_label,
        round(metrics.get("mrr",        0), 4),
        round(metrics.get("hits_at_1",  0), 4),
        round(metrics.get("hits_at_3",  0), 4),
        round(metrics.get("hits_at_10", 0), 4),
        round(metrics.get("mean_rank",  0), 4),
    ]

    for col, val in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col, value=val)

    wb.save(filepath)
    print(f"  Recorded: {filepath.name}  Run #{run_num}")


# ══════════════════════════════════════════════════════════════════════════════
# Main
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":

    print("=" * 60)
    print("Logistic Regression Baseline")
    print(f"  Source embeddings : {SOURCE_MODEL}")
    print(f"  Embedding file    : {EMB_PATH}")
    print("=" * 60)

    # ── Load embeddings ───────────────────────────────────────────────────────
    if not EMB_PATH.exists():
        raise FileNotFoundError(
            f"\nEmbedding file not found: {EMB_PATH}\n"
            f"Run the {SOURCE_MODEL} training script first to generate embeddings.\n"
            f"Expected file: {{model}}_sup_output/entity_embeddings.csv"
        )

    df_emb = pd.read_csv(EMB_PATH)
    print(f"Loaded {len(df_emb):,} entity embeddings")

    emb_cols = [c for c in df_emb.columns if c not in ('entity_id', 'entity_name', 'x', 'y')]

    if emb_cols:

        print(f"Using full embeddings  : {len(emb_cols)} dimensions")
        emb_lookup = {
            row['entity_name']: row[emb_cols].values.astype(np.float32)
            for _, row in df_emb.iterrows()
        }
        EMB_DIM = len(emb_cols)
    else:

        print("WARNING: Only 2D PCA coordinates found in entity_embeddings.csv.")
        print("         Results will be weaker than using full embeddings.")
        print("         Add full embedding export to your KGE script for best results.")
        emb_lookup = {
            row['entity_name']: np.array([row['x'], row['y']], dtype=np.float32)
            for _, row in df_emb.iterrows()
        }
        EMB_DIM = 2

    # ── Load and split data ───────────────────────────────────────────────────
    print("\nLoading triples...")
    df = pd.read_csv(CSV_PATH)
    df = df.drop_duplicates(subset=['head', 'relation', 'tail']).reset_index(drop=True)
    print(f"Clean triples: {len(df):,}")
    print(f"Relations    : {df['relation'].value_counts().to_dict()}")

    df_train, df_val, df_test = stratified_split(
        df, train_ratio=TRAIN_RATIO, val_ratio=VAL_RATIO, random_seed=RANDOM_SEED
    )
    # Combine train+val for final classifier training 
    df_trainval = pd.concat([df_train, df_val]).reset_index(drop=True)

    print(f"\nTrain+Val : {len(df_trainval):,}")
    print(f"Test      : {len(df_test):,}")

    # ── Baselines: majority class and random ──────────────────────────────────
    print("\n" + "=" * 60)
    print("Baselines")
    print("=" * 60)

    rel_counts  = df_test['relation'].value_counts()
    majority    = rel_counts.idxmax()
    n_test      = len(df_test)
    relations   = sorted(df_test['relation'].unique())

    print(f"\nMajority class: '{majority}'  ({rel_counts[majority]:,} / {n_test:,} = "
          f"{rel_counts[majority]/n_test:.1%})")

    y_majority = np.array([majority] * n_test)
    y_random   = np.random.RandomState(RANDOM_SEED).choice(relations, size=n_test)
    y_true_all = df_test['relation'].values

    for label, y_pred in [("Majority class", y_majority), ("Random (uniform)", y_random)]:
        df_per, macro = compute_per_relation_metrics(y_true_all, y_pred, relations)
        print(f"\n  {label}:")
        print(f"    {'Relation':<20} {'Count':>6} {'Hits@1':>8}")
        print(f"    {'-'*38}")
        for rel, row in df_per.iterrows():
            print(f"    {rel:<20} {row['count']:>6} {row['hits_at_1']:>8.4f}")
        print(f"    {'-'*38}")
        print(f"    {'Macro-Hits@1':<20} {'':>6} {macro:>8.4f}")

    # ── Logistic Regression — all feature strategies ──────────────────────────
    print("\n" + "=" * 60)
    print("Logistic Regression")
    print(f"  Embeddings from : {SOURCE_MODEL.upper()}")
    print(f"  Embedding dim   : {EMB_DIM}")
    print(f"  C               : {C}")
    print(f"  class_weight    : balanced")
    print(f"  Feature strategies: {FEATURE_STRATEGIES}")
    print("=" * 60)

    best_macro   = -1.0
    best_strategy = None
    all_results  = {}

    for strategy in FEATURE_STRATEGIES:
        print(f"\n--- Strategy: {strategy} ---")

        # Build features
        X_train, y_train, _         = build_features(df_trainval, emb_lookup, strategy)
        X_test,  y_test,  valid_idx = build_features(df_test,     emb_lookup, strategy)

        if len(X_train) == 0 or len(X_test) == 0:
            print("  [SKIP] No valid triples after embedding lookup.")
            continue

        feat_dim = X_train.shape[1]
        print(f"  Train samples : {len(X_train):,}  |  feature dim : {feat_dim}")
        print(f"  Test samples  : {len(X_test):,}")

        # Scale features — important for LR convergence
        scaler  = StandardScaler()
        X_train = scaler.fit_transform(X_train)
        X_test  = scaler.transform(X_test)

        # Train
        t0 = time.time()
        clf = LogisticRegression(
            C            = C,
            max_iter     = MAX_ITER,
            class_weight = "balanced",   # compensates for NOT_CONTRARY majority
            random_state = RANDOM_SEED,
            solver       = "lbfgs",
        )
        clf.fit(X_train, y_train)
        elapsed = time.time() - t0

        # Predict
        y_pred = clf.predict(X_test)

        # ── Save per-triple evaluation results ───────────────────────────────────
        df_eval_triples = df_test.loc[valid_idx].copy().reset_index(drop=True)

        df_eval_triples['predicted']  = y_pred
        df_eval_triples['true_label'] = y_test
        df_eval_triples['correct']    = (y_pred == y_test)

        # Save one CSV per strategy so you can compare them
        eval_path = ROOT / "outputs" / f"{SOURCE_MODEL}_sup_output" / f"lr_{strategy}_eval.csv"
        df_eval_triples[['head', 'relation', 'tail', 'domain',
                        'true_label', 'predicted', 'correct']].to_csv(eval_path, index=False)
        print(f"  Saved: {eval_path.name}  ({len(df_eval_triples):,} triples)")
        # Per-relation metrics
        df_per, macro = compute_per_relation_metrics(y_test, y_pred, relations)

        # Micro accuracy
        micro = (y_pred == y_test).mean()

        print(f"\n  {'Relation':<20} {'Count':>6} {'Correct':>8} {'Hits@1':>8}")
        print(f"  {'-'*46}")
        for rel, row in df_per.iterrows():
            print(f"  {rel:<20} {row['count']:>6} {row['correct']:>8} {row['hits_at_1']:>8.4f}")
        print(f"  {'-'*46}")
        print(f"  {'Macro-Hits@1':<20} {'':>6} {'':>8} {macro:>8.4f}  ← primary metric")
        print(f"  {'Micro-Hits@1':<20} {'':>6} {'':>8} {micro:>8.4f}")
        print(f"  Training time : {elapsed:.2f}s")

        # Confusion matrix
        cm = confusion_matrix(y_test, y_pred, labels=relations)
        print(f"\n  Confusion matrix (rows=true, cols=pred):")
        print(f"  Labels: {relations}")
        for i, row in enumerate(cm):
            print(f"    {relations[i]:<20}: {row}")

        all_results[strategy] = {
            "macro":   macro,
            "micro":   micro,
            "elapsed": elapsed,
            "df_per":  df_per,
        }

        if macro > best_macro:
            best_macro    = macro
            best_strategy = strategy

        # Record to Excel
        settings = {
            "dataset":       "ALL",
            "model":         f"LR-{SOURCE_MODEL}-{strategy}",
            "loss_func":     "LogLoss",
            "neg_sampler":   "—",
            "corrupt":       "—",
            "embedding_dim": EMB_DIM,
            "num_epochs":    "—",
            "batch_size":    "—",
            "learning_rate": f"C={C}",
            "margin":        "—",
            "adv_temp":      "—",
            "num_neg":       "—",
        }
        metrics = {
            "mrr":        macro,  
            "hits_at_1":  micro,
            "hits_at_3":  1.0,     
            "hits_at_10": 1.0,
            "mean_rank":  0.0,     # not applicable for LR
        }
        record_results(
            RESULTS_FILE, settings, metrics, elapsed,
            eval_label=f"LR-{strategy}"
        )

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("Summary — all strategies")
    print("=" * 60)
    print(f"\n  {'Strategy':<12} {'Macro-H@1':>10} {'Micro-H@1':>10} {'Time':>8}")
    print(f"  {'-'*44}")
    for strat, res in all_results.items():
        marker = " <-- best" if strat == best_strategy else ""
        print(f"  {strat:<12} {res['macro']:>10.4f} {res['micro']:>10.4f} "
              f"{res['elapsed']:>7.2f}s{marker}")
    print(f"\n  Best strategy : {best_strategy}  (Macro-Hits@1 = {best_macro:.4f})")

    # ── LR vs KGE comparison ──────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("LR vs KGE Comparison  (Macro-Hits@1)")
    print("=" * 60)

    try:
        df_xl = pd.read_excel(RESULTS_FILE)

        kge_rows = df_xl[df_xl["Evaluator"].str.endswith("-Macro", na=False)][
            ["Model", "MRR"]
        ].copy()
        kge_rows.columns = ["Model", "Macro-H@1"]
        kge_rows = kge_rows.drop_duplicates("Model").sort_values("Macro-H@1", ascending=False)

        lr_rows = df_xl[df_xl["Evaluator"].str.startswith("LR-", na=False)][
            ["Model", "MRR"]
        ].copy()
        lr_rows.columns = ["Model", "Macro-H@1"]
        # Keep only rows from the current SOURCE_MODEL run
        lr_rows = lr_rows[lr_rows["Model"].str.startswith(f"LR-{SOURCE_MODEL}-")]
        lr_rows = lr_rows.drop_duplicates("Model").sort_values("Macro-H@1", ascending=False)

        best_lr_macro = lr_rows["Macro-H@1"].max() if len(lr_rows) else None

        print(f"\n  KGE models  (Macro-H@1):")
        print(f"  {'Model':<20} {'Macro-H@1':>10}  vs LR-best")
        print(f"  {'-'*44}")
        for _, row in kge_rows.iterrows():
            if best_lr_macro is not None:
                diff   = row["Macro-H@1"] - best_lr_macro
                marker = f"  {'KGE better' if diff > 0.001 else ('≈ tied' if abs(diff) <= 0.001 else 'LR better')}  ({diff:+.4f})"
            else:
                marker = ""
            print(f"  {row['Model']:<20} {row['Macro-H@1']:>10.4f}{marker}")

        print(f"\n  LR baseline  (embeddings from {SOURCE_MODEL.upper()}):")
        print(f"  {'Model':<30} {'Macro-H@1':>10}")
        print(f"  {'-'*44}")
        for _, row in lr_rows.iterrows():
            marker = "  <-- best" if row["Macro-H@1"] == best_lr_macro else ""
            print(f"  {row['Model']:<30} {row['Macro-H@1']:>10.4f}{marker}")

        if best_lr_macro is not None and len(kge_rows):
            best_kge_macro = kge_rows["Macro-H@1"].max()
            gap = best_kge_macro - best_lr_macro
            print(f"\n  Best KGE : {best_kge_macro:.4f}  |  Best LR : {best_lr_macro:.4f}  |  Gap : {gap:+.4f}")

    except Exception as e:
        print(f"  [Could not load comparison from Excel: {e}]")

    print("\n" + "=" * 60)
    print("All done.")
    print("=" * 60)