"""
Comprehensive unit test suite
==============================
Covers all 5 sup_* KGE models, lr_baseline, lr_onehot, and caliensemble.

Coverage map
────────────
SUP MODELS  (RotatE · ComplEx · DistMult · ConvKB · TransE)  — per model:
  A  stratified_split
  B  ExtendedBasicNegativeSampler
  C  RelationRankEvaluator
  D  record_results
  E  get_node_level
  F  build_aba_tree

LR BASELINE
  G  build_features  (concat / diff / hadamard / all)
  H  compute_per_relation_metrics
  I  record_results  (eval_label="LR")

LR ONEHOT
  J  compute_per_relation_metrics
  K  record_results  (eval_label="OneHot-LR")

CALIENSEMBLE
  L  fit_calibrators
  M  apply_calibrators
  N  compute_metrics

Claude-Assisted
"""

import importlib
import sys
import tempfile
import types
import unittest
import datetime
from pathlib import Path
from unittest.mock import MagicMock

import numpy as np
import pandas as pd
import torch
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split

# ─────────────────────────────────────────────────────────────────────────────
# Stub pykeen BEFORE importing any sup_* module
# ─────────────────────────────────────────────────────────────────────────────
_pykeen          = types.ModuleType("pykeen")
_pykeen_pipeline = types.ModuleType("pykeen.pipeline")
_pykeen_triples  = types.ModuleType("pykeen.triples")
_pykeen_sampling = types.ModuleType("pykeen.sampling")

class _StubBasicNegativeSampler(torch.nn.Module):
    num_negs_per_pos  = 1
    corruption_scheme = ["relation"]
    def __init__(self, *args, **kwargs): super().__init__()
    def corrupt_batch(self, positive_batch): return positive_batch.clone()

_pykeen_sampling.BasicNegativeSampler = _StubBasicNegativeSampler
_pykeen_pipeline.pipeline             = MagicMock()
_pykeen_triples.TriplesFactory        = MagicMock()
for _n, _m in [
    ("pykeen", _pykeen), ("pykeen.pipeline", _pykeen_pipeline),
    ("pykeen.triples", _pykeen_triples), ("pykeen.sampling", _pykeen_sampling),
]:
    sys.modules.setdefault(_n, _m)

# ─────────────────────────────────────────────────────────────────────────────
# Path setup and module imports
# ─────────────────────────────────────────────────────────────────────────────
sys.path.insert(0, str(Path(__file__).parent))

_SUP_TAGS = {
    "sup_rotate":   "RotatE",
    "sup_complex":  "ComplEx",
    "sup_distmult": "DistMult",
    "sup_convkb":   "ConvKB",
    "sup_transe":   "TransE",
}
_sup_mods = {name: importlib.import_module(name) for name in _SUP_TAGS}

import lr_baseline as _lr
import lr_onehot   as _oh
import caliensemble as _cali

# ─────────────────────────────────────────────────────────────────────────────
# Shared test helpers
# ─────────────────────────────────────────────────────────────────────────────

def _large_df(n=200, domains=("location", "cleanliness"),
              rels=("CONTRARY_TO", "SUPPORT")):
    return pd.DataFrame([
        {"head": f"phrase_{i % 40}", "tail": f"phrase_{(i+1) % 40}",
         "relation": rels[i % len(rels)], "domain": domains[i % len(domains)]}
        for i in range(n)
    ])


def _make_tf(num_entities=6, num_relations=3):
    tf = MagicMock()
    tf.num_entities   = num_entities
    tf.num_relations  = num_relations
    tf.entity_to_id   = {f"e{i}": i for i in range(num_entities)}
    tf.relation_to_id = {f"r{i}": i for i in range(num_relations)}
    tf.mapped_triples = torch.tensor(
        [[i % num_entities, i % num_relations, (i + 2) % num_entities]
         for i in range(12)],
        dtype=torch.long,
    )
    return tf


def _const_model(score_val=1.0):
    m = MagicMock()
    m.score_hrt = MagicMock(
        side_effect=lambda hrt: torch.full((hrt.shape[0], 1), float(score_val))
    )
    return m


def _make_settings(**ov):
    d = dict(dataset="TEST", model="RotatE", loss_func="NSSALoss",
             neg_sampler="ExtBNS", corrupt="relation", embedding_dim=50,
             num_epochs=100, best_epoch=40, batch_size=128,
             learning_rate=0.001, margin=1.0, adv_temp=0.5, num_neg=2)
    d.update(ov)
    return d


def _make_metrics(**ov):
    d = dict(mrr=0.80, hits_at_1=0.65, hits_at_3=0.85,
             hits_at_10=0.95, mean_rank=1.3)
    d.update(ov)
    return d


def _aba_df():
    return pd.DataFrame([
        {"head": "nice_rooms",    "relation": "SUPPORT",     "tail": "good_staff",    "score": 0.9, "domain": "rooms"},
        {"head": "clean_lobby",   "relation": "SUPPORT",     "tail": "good_price",    "score": 0.8, "domain": "lobby"},
        {"head": "fast_checkin",  "relation": "SUPPORT",     "tail": "good_check-in", "score": 0.7, "domain": "checkin"},
        {"head": "no_evident_rooms", "relation": "CONTRARY_TO", "tail": "nice_rooms", "score": 0.3, "domain": "rooms"},
        {"head": "dirty_lobby",   "relation": "NOT_CONTRARY","tail": "clean_lobby",   "score": 0.2, "domain": "lobby"},
        {"head": "fast_checkin",  "relation": "SUPPORT",     "tail": "good_price",    "score": 0.6, "domain": "checkin"},
        {"head": "att1",          "relation": "CONTRARY_TO", "tail": "fast_checkin",  "score": 0.1, "domain": "checkin"},
        {"head": "att2",          "relation": "NOT_CONTRARY","tail": "fast_checkin",  "score": 0.15,"domain": "checkin"},
    ])


def _b(*rows):
    return torch.tensor(list(rows), dtype=torch.long)


# ─────────────────────────────────────────────────────────────────────────────
# Factory: generate one full test suite per sup_* model
# ─────────────────────────────────────────────────────────────────────────────

def _make_sup_suite(mod_name, tag):
    """Return a list of TestCase classes for one sup model."""
    mod    = _sup_mods[mod_name]
    CLAIMS = mod.CLAIMS
    EMAP   = {0: "cat", 1: "dog", 2: "fish", 3: "bird"}
    RMAP   = {0: "CONTRARY_TO", 1: "SUPPORT", 2: "NOT_CONTRARY"}

    def _sampler(td):
        mod.id_to_entity   = EMAP
        mod.id_to_relation = RMAP
        return mod.ExtendedBasicNegativeSampler(save_path=str(Path(td) / "neg.csv"))

    def _ev(score_val=1.0, num_rel=3):
        tf = _make_tf(num_relations=num_rel)
        mod.id_to_entity   = {v: k for k, v in tf.entity_to_id.items()}
        mod.id_to_relation = {v: k for k, v in tf.relation_to_id.items()}
        ev = mod.RelationRankEvaluator(_const_model(score_val), tf,
                                       batch_size=4, device="cpu")
        return ev, tf

    def _fp(td):
        return str(Path(td) / "r.xlsx")

    classes = []

    # ── A: stratified_split ───────────────────────────────────────────────────
    class A(unittest.TestCase):
        def _split(self, n=200):
            df = _large_df(n)
            df["_i"] = range(len(df))
            t, v, te = mod.stratified_split(df)
            return set(t["_i"]), set(v["_i"]), set(te["_i"])

        def test_no_leakage_train_val(self):
            ti, vi, _ = self._split()
            self.assertEqual(ti & vi, set())

        def test_no_leakage_train_test(self):
            ti, _, tei = self._split()
            self.assertEqual(ti & tei, set())

        def test_no_leakage_val_test(self):
            _, vi, tei = self._split()
            self.assertEqual(vi & tei, set())

        def test_total_rows_preserved(self):
            df = _large_df()
            t, v, te = mod.stratified_split(df)
            self.assertEqual(len(t) + len(v) + len(te), len(df))

        def test_train_ratio_approx_80(self):
            df = _large_df(400)
            t, v, te = mod.stratified_split(df, train_ratio=0.8, val_ratio=0.1)
            self.assertAlmostEqual(len(t) / len(df), 0.8, delta=0.05)

        def test_val_ratio_approx_10(self):
            df = _large_df(400)
            t, v, te = mod.stratified_split(df, train_ratio=0.8, val_ratio=0.1)
            self.assertAlmostEqual(len(v) / len(df), 0.1, delta=0.05)

        def test_returns_three_dataframes(self):
            splits = mod.stratified_split(_large_df(120))
            self.assertEqual(len(splits), 3)
            for s in splits:
                self.assertIsInstance(s, pd.DataFrame)

        def test_reproducible_same_seed(self):
            df = _large_df(120)
            t1, _, _ = mod.stratified_split(df, random_seed=7)
            t2, _, _ = mod.stratified_split(df, random_seed=7)
            pd.testing.assert_frame_equal(
                t1.reset_index(drop=True), t2.reset_index(drop=True))

        def test_different_seeds_differ(self):
            df = _large_df(200)
            t1, _, _ = mod.stratified_split(df, random_seed=1)
            t2, _, _ = mod.stratified_split(df, random_seed=99)
            self.assertFalse((t1["head"].values == t2["head"].values).all())

        def test_index_reset_in_splits(self):
            for split in mod.stratified_split(_large_df(120)):
                if len(split) > 0:
                    self.assertEqual(list(split.index), list(range(len(split))))

    A.__name__     = f"Test_{tag}_StratifiedSplit"
    A.__qualname__ = A.__name__
    classes.append(A)

    # ── B: ExtendedBasicNegativeSampler ───────────────────────────────────────
    class B(unittest.TestCase):
        def test_corruption_scheme_forced_to_relation(self):
            with tempfile.TemporaryDirectory() as td:
                self.assertEqual(_sampler(td).corruption_scheme, ["relation"])

        def test_caller_scheme_overridden(self):
            with tempfile.TemporaryDirectory() as td:
                mod.id_to_entity   = EMAP
                mod.id_to_relation = RMAP
                s = mod.ExtendedBasicNegativeSampler(
                    save_path=str(Path(td) / "neg.csv"),
                    corruption_scheme=["head"])
                self.assertEqual(s.corruption_scheme, ["relation"])

        def test_initial_batch_count_zero(self):
            with tempfile.TemporaryDirectory() as td:
                self.assertEqual(_sampler(td).batch_count, 0)

        def test_initial_header_written_false(self):
            with tempfile.TemporaryDirectory() as td:
                self.assertFalse(_sampler(td).header_written)

        def test_returns_long_tensor(self):
            with tempfile.TemporaryDirectory() as td:
                self.assertEqual(
                    _sampler(td).corrupt_batch(_b([0, 1, 2])).dtype, torch.long)

        def test_returns_same_shape(self):
            with tempfile.TemporaryDirectory() as td:
                b = _b([0, 1, 2], [1, 0, 3])
                self.assertEqual(_sampler(td).corrupt_batch(b).shape, b.shape)

        def test_batch_count_increments(self):
            with tempfile.TemporaryDirectory() as td:
                s = _sampler(td)
                for i in range(1, 5):
                    s.corrupt_batch(positive_batch=_b([0, 1, 2]))
                    self.assertEqual(s.batch_count, i)

        def test_csv_not_created_before_first_call(self):
            with tempfile.TemporaryDirectory() as td:
                _sampler(td)
                self.assertFalse((Path(td) / "neg.csv").exists())

        def test_csv_created_after_first_call(self):
            with tempfile.TemporaryDirectory() as td:
                s = _sampler(td)
                s.corrupt_batch(_b([0, 1, 2]))
                self.assertTrue((Path(td) / "neg.csv").exists())

        def test_header_written_flag_updates(self):
            with tempfile.TemporaryDirectory() as td:
                s = _sampler(td)
                self.assertFalse(s.header_written)
                s.corrupt_batch(_b([0, 1, 2]))
                self.assertTrue(s.header_written)

        def test_csv_columns(self):
            with tempfile.TemporaryDirectory() as td:
                s = _sampler(td)
                s.corrupt_batch(_b([0, 1, 2]))
                self.assertEqual(
                    set(pd.read_csv(Path(td) / "neg.csv").columns),
                    {"head", "relation", "tail", "label"})

        def test_all_labels_zero(self):
            with tempfile.TemporaryDirectory() as td:
                s = _sampler(td)
                s.corrupt_batch(_b([0, 1, 2], [3, 0, 1]))
                self.assertTrue(
                    (pd.read_csv(Path(td) / "neg.csv")["label"] == 0).all())

        def test_multiple_batches_accumulate(self):
            with tempfile.TemporaryDirectory() as td:
                s = _sampler(td)
                for _ in range(3):
                    s.corrupt_batch(_b([0, 1, 2], [1, 2, 3]))
                self.assertEqual(len(pd.read_csv(Path(td) / "neg.csv")), 6)

        def test_header_written_exactly_once(self):
            with tempfile.TemporaryDirectory() as td:
                s = _sampler(td)
                for _ in range(4):
                    s.corrupt_batch(_b([0, 1, 2]))
                lines = open(Path(td) / "neg.csv").readlines()
                self.assertEqual(
                    sum(1 for l in lines if l.startswith("head,")), 1)

        def test_unknown_id_maps_to_unknown(self):
            with tempfile.TemporaryDirectory() as td:
                s = _sampler(td)
                s.corrupt_batch(_b([99, 99, 99]))
                df = pd.read_csv(Path(td) / "neg.csv")
                self.assertTrue((df["head"] == "unknown").all())

    B.__name__     = f"Test_{tag}_NegSampler"
    B.__qualname__ = B.__name__
    classes.append(B)

    # ── C: RelationRankEvaluator ──────────────────────────────────────────────
    class C(unittest.TestCase):
        def test_returns_tuple_of_two(self):
            ev, tf = _ev()
            self.assertEqual(len(ev.evaluate(tf.mapped_triples)), 2)

        def test_first_is_dict(self):
            ev, tf = _ev()
            m, _ = ev.evaluate(tf.mapped_triples)
            self.assertIsInstance(m, dict)

        def test_second_is_dataframe(self):
            ev, tf = _ev()
            _, df = ev.evaluate(tf.mapped_triples)
            self.assertIsInstance(df, pd.DataFrame)

        def test_pykeen_compatible_keys(self):
            ev, tf = _ev()
            m, _ = ev.evaluate(tf.mapped_triples)
            r = m["both"]["realistic"]
            for k in ("inverse_harmonic_mean_rank", "arithmetic_mean_rank",
                      "hits_at_1", "hits_at_3", "hits_at_10"):
                self.assertIn(k, r)

        def test_row_count_equals_triple_count(self):
            ev, tf = _ev()
            _, df = ev.evaluate(tf.mapped_triples)
            self.assertEqual(len(df), len(tf.mapped_triples))

        def test_required_df_columns(self):
            ev, tf = _ev()
            _, df = ev.evaluate(tf.mapped_triples)
            for col in ("head", "relation", "tail", "relation_rank",
                        "reciprocal_rank", "hits_at_1", "hits_at_3", "hits_at_10"):
                self.assertIn(col, df.columns)

        def test_rank_always_gte_1(self):
            ev, tf = _ev()
            _, df = ev.evaluate(tf.mapped_triples)
            self.assertTrue((df["relation_rank"] >= 1).all())

        def test_reciprocal_rank_in_0_to_1(self):
            ev, tf = _ev()
            _, df = ev.evaluate(tf.mapped_triples)
            self.assertTrue((df["reciprocal_rank"] > 0).all())
            self.assertTrue((df["reciprocal_rank"] <= 1.0).all())

        def test_hits_monotone_per_row(self):
            ev, tf = _ev()
            _, df = ev.evaluate(tf.mapped_triples)
            self.assertTrue((df["hits_at_10"] >= df["hits_at_3"]).all())
            self.assertTrue((df["hits_at_3"]  >= df["hits_at_1"]).all())

        def test_mrr_equals_mean_rr(self):
            ev, tf = _ev()
            m, df = ev.evaluate(tf.mapped_triples)
            self.assertAlmostEqual(
                m["both"]["realistic"]["inverse_harmonic_mean_rank"],
                round(df["reciprocal_rank"].mean(), 4), places=3)

        def test_mean_rank_equals_mean_rel_rank(self):
            ev, tf = _ev()
            m, df = ev.evaluate(tf.mapped_triples)
            self.assertAlmostEqual(
                m["both"]["realistic"]["arithmetic_mean_rank"],
                round(df["relation_rank"].mean(), 4), places=3)

        def test_perfect_score_rank_one(self):
            tf = _make_tf(num_relations=3)
            mod.id_to_entity   = {v: k for k, v in tf.entity_to_id.items()}
            mod.id_to_relation = {v: k for k, v in tf.relation_to_id.items()}
            mdl = MagicMock()
            def _s(hrt):
                s = torch.zeros(hrt.shape[0], 1)
                for i, r in enumerate(hrt):
                    if r[1].item() == 0: s[i] = 10.0
                return s
            mdl.score_hrt = _s
            ev = mod.RelationRankEvaluator(mdl, tf, device="cpu")
            _, df = ev.evaluate(torch.tensor([[0, 0, 2]], dtype=torch.long))
            self.assertEqual(df.iloc[0]["relation_rank"], 1)

        def test_worst_score_rank_num_relations(self):
            tf = _make_tf(num_relations=3)
            mod.id_to_entity   = {v: k for k, v in tf.entity_to_id.items()}
            mod.id_to_relation = {v: k for k, v in tf.relation_to_id.items()}
            mdl = MagicMock()
            def _s(hrt):
                s = torch.full((hrt.shape[0], 1), 10.0)
                for i, r in enumerate(hrt):
                    if r[1].item() == 0: s[i] = 0.0
                return s
            mdl.score_hrt = _s
            ev = mod.RelationRankEvaluator(mdl, tf, device="cpu")
            _, df = ev.evaluate(torch.tensor([[0, 0, 2]], dtype=torch.long))
            self.assertEqual(df.iloc[0]["relation_rank"], 3)

    C.__name__     = f"Test_{tag}_RelationRankEvaluator"
    C.__qualname__ = C.__name__
    classes.append(C)

    # ── D: record_results ─────────────────────────────────────────────────────
    class D(unittest.TestCase):
        def test_creates_file(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                self.assertTrue(Path(_fp(td)).exists())

        def test_header_row_no_col_1(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                ws = load_workbook(_fp(td)).active
                self.assertEqual(ws.cell(1, 1).value, "No.")

        def test_header_mrr_col_18(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                ws = load_workbook(_fp(td)).active
                self.assertEqual(ws.cell(1, 18).value, "MRR")

        def test_freeze_pane_A2(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                self.assertEqual(
                    str(load_workbook(_fp(td)).active.freeze_panes), "A2")

        def test_sheet_title(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                self.assertEqual(
                    load_workbook(_fp(td)).active.title, "Experiment Results")

        def test_first_run_number_1(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                self.assertEqual(
                    load_workbook(_fp(td)).active.cell(2, 1).value, 1)

        def test_second_call_appends(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                self.assertEqual(
                    load_workbook(_fp(td)).active.max_row, 3)

        def test_run_numbers_increment(self):
            with tempfile.TemporaryDirectory() as td:
                for _ in range(4):
                    mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                ws = load_workbook(_fp(td)).active
                self.assertEqual(
                    [ws.cell(r, 1).value for r in range(2, 6)], [1, 2, 3, 4])

        def test_mrr_col_18(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(mrr=0.77), 0)
                self.assertAlmostEqual(
                    load_workbook(_fp(td)).active.cell(2, 18).value, 0.77)

        def test_elapsed_3661s(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 3661)
                self.assertEqual(
                    load_workbook(_fp(td)).active.cell(2, 16).value, "1h 1m 1s")

        def test_elapsed_0s(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                self.assertEqual(
                    load_workbook(_fp(td)).active.cell(2, 16).value, "0h 0m 0s")

        def test_eval_label_col_17(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(
                    _fp(td), _make_settings(), _make_metrics(), 0, "RelRank")
                self.assertEqual(
                    load_workbook(_fp(td)).active.cell(2, 17).value, "RelRank")

        def test_metrics_rounded_4_places(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(
                    _fp(td), _make_settings(), _make_metrics(mrr=0.123456789), 0)
                self.assertEqual(
                    load_workbook(_fp(td)).active.cell(2, 18).value, 0.1235)

        def test_missing_best_epoch_defaults_dash(self):
            with tempfile.TemporaryDirectory() as td:
                s = _make_settings()
                del s["best_epoch"]
                mod.record_results(_fp(td), s, _make_metrics(), 0)
                self.assertEqual(
                    load_workbook(_fp(td)).active.cell(2, 10).value, "—")

        def test_timestamp_format(self):
            with tempfile.TemporaryDirectory() as td:
                mod.record_results(_fp(td), _make_settings(), _make_metrics(), 0)
                ts = load_workbook(_fp(td)).active.cell(2, 2).value
                datetime.datetime.strptime(ts, "%Y-%m-%d %H:%M")

    D.__name__     = f"Test_{tag}_RecordResults"
    D.__qualname__ = D.__name__
    classes.append(D)

    # ── E: get_node_level ─────────────────────────────────────────────────────
    class E(unittest.TestCase):
        def test_all_claims_return_claim(self):
            for c in CLAIMS:
                self.assertEqual(mod.get_node_level(c), "claim")

        def test_no_evident_prefix_is_attacker(self):
            for sfx in ("staff", "price", "rooms", "x"):
                self.assertEqual(mod.get_node_level(f"no_evident_{sfx}"), "attacker")

        def test_regular_phrases_are_body(self):
            for n in ("nice_rooms", "fast_checkin", "clean_lobby", ""):
                self.assertEqual(mod.get_node_level(n), "body")

        def test_claim_not_attacker(self):
            for c in CLAIMS:
                self.assertNotEqual(mod.get_node_level(c), "attacker")

        def test_claim_not_body(self):
            for c in CLAIMS:
                self.assertNotEqual(mod.get_node_level(c), "body")

        def test_no_evident_not_body(self):
            self.assertNotEqual(mod.get_node_level("no_evident_staff"), "body")

        def test_no_evident_not_claim(self):
            self.assertNotEqual(mod.get_node_level("no_evident_staff"), "claim")

        def test_no_evident_with_claim_suffix_still_attacker(self):
            self.assertEqual(mod.get_node_level("no_evident_good_staff"), "attacker")

        def test_typo_claim_is_body(self):
            self.assertEqual(mod.get_node_level("good_staf"), "body")

    E.__name__     = f"Test_{tag}_GetNodeLevel"
    E.__qualname__ = E.__name__
    classes.append(E)

    # ── F: build_aba_tree ─────────────────────────────────────────────────────
    class F(unittest.TestCase):
        def test_returns_dataframe(self):
            self.assertIsInstance(mod.build_aba_tree(_aba_df(), CLAIMS), pd.DataFrame)

        def test_required_columns(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            for col in ("claim", "body", "attacker", "body_relation",
                        "attack_relation", "body_score", "attack_score",
                        "domain", "body_level"):
                self.assertIn(col, df.columns)

        def test_support_edges_create_bodies(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            self.assertIn("nice_rooms", df["body"].values)

        def test_claim_column_only_valid_claims(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            self.assertTrue(set(df["claim"].unique()).issubset(set(CLAIMS)))

        def test_attacker_present_for_attacked_body(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            rows = df[(df["body"] == "nice_rooms") & df["attacker"].notna()]
            self.assertGreater(len(rows), 0)

        def test_attacker_none_for_unattacked_body(self):
            simple = pd.DataFrame([
                {"head": "fast_checkin", "relation": "SUPPORT",
                 "tail": "good_check-in", "score": 0.7, "domain": "checkin"}
            ])
            result = mod.build_aba_tree(simple, CLAIMS)
            self.assertTrue(result["attacker"].isna().all())

        def test_body_two_attackers_two_rows(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            rows = df[(df["claim"] == "good_price") & (df["body"] == "fast_checkin")]
            self.assertEqual(len(rows), 2)

        def test_body_relation_always_support(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            self.assertTrue((df["body_relation"] == "SUPPORT").all())

        def test_attack_relation_valid_types(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            non_null = df[df["attack_relation"].notna()]
            self.assertTrue(
                non_null["attack_relation"].isin({"CONTRARY_TO", "NOT_CONTRARY"}).all())

        def test_body_score_correct(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            self.assertAlmostEqual(
                df[df["body"] == "nice_rooms"].iloc[0]["body_score"], 0.9)

        def test_attack_score_none_when_no_attacker(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            self.assertTrue(df[df["attacker"].isna()]["attack_score"].isna().all())

        def test_domain_propagated(self):
            df = mod.build_aba_tree(_aba_df(), CLAIMS)
            self.assertFalse(df["domain"].isna().all())

        def test_empty_input_returns_empty(self):
            empty = pd.DataFrame(
                columns=["head", "relation", "tail", "score", "domain"])
            self.assertEqual(len(mod.build_aba_tree(empty, CLAIMS)), 0)

    F.__name__     = f"Test_{tag}_BuildABATree"
    F.__qualname__ = F.__name__
    classes.append(F)

    return classes


# ─────────────────────────────────────────────────────────────────────────────
# Register all 5 sup model test suites into the module namespace
# ─────────────────────────────────────────────────────────────────────────────
for _mod_name, _tag in _SUP_TAGS.items():
    for _cls in _make_sup_suite(_mod_name, _tag):
        globals()[_cls.__name__] = _cls


# =============================================================================
# G  lr_baseline — build_features
# =============================================================================
class TestLR_BuildFeatures(unittest.TestCase):

    D = 4  # embedding dimension

    def _emb(self, n=5):
        rng = np.random.default_rng(0)
        return {f"e{i}": rng.random(self.D).astype(np.float32) for i in range(n)}

    def _df(self, n=5):
        return pd.DataFrame([
            {"head": f"e{i}", "tail": f"e{(i+1) % 5}",
             "relation": "A" if i % 2 == 0 else "B"}
            for i in range(n)
        ])

    def test_concat_shape(self):
        X, y, idx = _lr.build_features(self._df(), self._emb(), "concat")
        self.assertEqual(X.shape, (5, self.D * 2))

    def test_diff_shape(self):
        X, y, idx = _lr.build_features(self._df(), self._emb(), "diff")
        self.assertEqual(X.shape, (5, self.D))

    def test_hadamard_shape(self):
        X, y, idx = _lr.build_features(self._df(), self._emb(), "hadamard")
        self.assertEqual(X.shape, (5, self.D))

    def test_all_shape(self):
        X, y, idx = _lr.build_features(self._df(), self._emb(), "all")
        self.assertEqual(X.shape, (5, self.D * 4))

    def test_dtype_float32(self):
        X, _, _ = _lr.build_features(self._df(), self._emb())
        self.assertEqual(X.dtype, np.float32)

    def test_default_strategy_is_concat(self):
        X_default, _, _ = _lr.build_features(self._df(), self._emb())
        X_concat,  _, _ = _lr.build_features(self._df(), self._emb(), "concat")
        np.testing.assert_array_equal(X_default, X_concat)

    def test_missing_entity_skipped(self):
        # only e0, e1, e2 in lookup — triples needing e3 or e4 get skipped
        emb = {f"e{i}": np.zeros(self.D, dtype=np.float32) for i in range(3)}
        X, y, idx = _lr.build_features(self._df(5), emb, "concat")
        self.assertLess(len(X), 5)

    def test_unknown_strategy_raises_value_error(self):
        with self.assertRaises(ValueError):
            _lr.build_features(self._df(), self._emb(), "UNKNOWN")

    def test_y_labels_match_relations(self):
        _, y, _ = _lr.build_features(self._df(), self._emb())
        self.assertEqual(set(y), {"A", "B"})

    def test_valid_indices_subset_of_original(self):
        df = self._df()
        _, _, idx = _lr.build_features(df, self._emb())
        self.assertTrue(set(idx).issubset(set(df.index)))

    def test_diff_is_h_minus_t(self):
        emb = {
            "h": np.array([2.0, 2.0], dtype=np.float32),
            "t": np.array([1.0, 1.0], dtype=np.float32),
        }
        df = pd.DataFrame([{"head": "h", "tail": "t", "relation": "R"}])
        X, _, _ = _lr.build_features(df, emb, "diff")
        np.testing.assert_allclose(X[0], np.array([1.0, 1.0]))

    def test_hadamard_is_element_wise(self):
        emb = {
            "h": np.array([2.0, 3.0], dtype=np.float32),
            "t": np.array([4.0, 5.0], dtype=np.float32),
        }
        df = pd.DataFrame([{"head": "h", "tail": "t", "relation": "R"}])
        X, _, _ = _lr.build_features(df, emb, "hadamard")
        np.testing.assert_allclose(X[0], np.array([8.0, 15.0]))

    def test_all_concatenates_four_parts(self):
        d = 2
        emb = {
            "h": np.array([1.0, 0.0], dtype=np.float32),
            "t": np.array([0.0, 1.0], dtype=np.float32),
        }
        df = pd.DataFrame([{"head": "h", "tail": "t", "relation": "R"}])
        X, _, _ = _lr.build_features(df, emb, "all")
        expected = np.array([1.0, 0.0, 0.0, 1.0, 1.0, -1.0, 0.0, 0.0],
                            dtype=np.float32)
        np.testing.assert_allclose(X[0], expected)

    def test_row_count_equals_valid_triples(self):
        emb = self._emb(5)
        df  = self._df(5)
        X, y, idx = _lr.build_features(df, emb)
        self.assertEqual(len(X), len(y))
        self.assertEqual(len(X), len(idx))


# =============================================================================
# H  lr_baseline — compute_per_relation_metrics
# =============================================================================
class TestLR_PerRelationMetrics(unittest.TestCase):

    def test_returns_df_and_scalar(self):
        y = np.array(["A", "A", "B", "B"])
        df, macro = _lr.compute_per_relation_metrics(y, y)
        self.assertIsInstance(df, pd.DataFrame)
        self.assertIsInstance(macro, float)

    def test_perfect_prediction_macro_1(self):
        y = np.array(["A", "A", "B", "B"])
        _, macro = _lr.compute_per_relation_metrics(y, y)
        self.assertAlmostEqual(macro, 1.0)

    def test_all_wrong_relation_hits_0(self):
        y_t = np.array(["A", "A"])
        y_p = np.array(["B", "B"])
        df, _ = _lr.compute_per_relation_metrics(y_t, y_p)
        self.assertEqual(df.loc["A", "hits_at_1"], 0.0)

    def test_index_is_relation_name(self):
        y = np.array(["A", "B", "C"])
        df, _ = _lr.compute_per_relation_metrics(y, y)
        self.assertEqual(set(df.index), {"A", "B", "C"})

    def test_count_column_correct(self):
        y_t = np.array(["A", "A", "A", "B"])
        y_p = np.array(["A", "A", "A", "B"])
        df, _ = _lr.compute_per_relation_metrics(y_t, y_p)
        self.assertEqual(df.loc["A", "count"], 3)

    def test_correct_column_correct(self):
        y_t = np.array(["A", "A", "B"])
        y_p = np.array(["A", "B", "B"])
        df, _ = _lr.compute_per_relation_metrics(y_t, y_p)
        self.assertEqual(df.loc["A", "correct"], 1)

    def test_macro_is_mean_of_per_relation(self):
        y_t = np.array(["A", "A", "B", "B"])
        y_p = np.array(["A", "B", "B", "A"])
        df, macro = _lr.compute_per_relation_metrics(y_t, y_p)
        self.assertAlmostEqual(macro, df["hits_at_1"].mean(), places=3)

    def test_hits_at_1_rounded_4_places(self):
        y_t = np.array(["A"] * 3 + ["B"] * 3)
        y_p = np.array(["A", "A", "B", "B", "B", "A"])
        df, _ = _lr.compute_per_relation_metrics(y_t, y_p)
        hits = df.loc["A", "hits_at_1"]
        self.assertEqual(hits, round(hits, 4))

    def test_custom_relations_unseen_class_hits_zero(self):
        y_t = np.array(["A", "B"])
        y_p = np.array(["A", "B"])
        df, _ = _lr.compute_per_relation_metrics(
            y_t, y_p, relations=["A", "B", "C"])
        self.assertIn("C", df.index)
        self.assertEqual(df.loc["C", "hits_at_1"], 0.0)


# =============================================================================
# I  lr_baseline — record_results
# =============================================================================
class TestLR_RecordResults(unittest.TestCase):

    def _fp(self, td): return Path(td) / "lr.xlsx"

    def test_creates_file(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            self.assertTrue(Path(self._fp(td)).exists())

    def test_default_eval_label_is_LR(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            ws = load_workbook(self._fp(td)).active
            self.assertEqual(ws.cell(2, 17).value, "LR")

    def test_header_no_col_1(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            ws = load_workbook(self._fp(td)).active
            self.assertEqual(ws.cell(1, 1).value, "No.")

    def test_header_mrr_col_18(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            ws = load_workbook(self._fp(td)).active
            self.assertEqual(ws.cell(1, 18).value, "MRR")

    def test_second_call_appends(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            _lr.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            self.assertEqual(load_workbook(self._fp(td)).active.max_row, 3)

    def test_mrr_stored_col_18(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(
                self._fp(td), _make_settings(), _make_metrics(mrr=0.66), 0)
            self.assertAlmostEqual(
                load_workbook(self._fp(td)).active.cell(2, 18).value, 0.66)

    def test_elapsed_formatted(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(
                self._fp(td), _make_settings(), _make_metrics(), 3661)
            self.assertEqual(
                load_workbook(self._fp(td)).active.cell(2, 16).value, "1h 1m 1s")

    def test_freeze_pane_A2(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            self.assertEqual(
                str(load_workbook(self._fp(td)).active.freeze_panes), "A2")

    def test_custom_eval_label(self):
        with tempfile.TemporaryDirectory() as td:
            _lr.record_results(
                self._fp(td), _make_settings(), _make_metrics(), 0, "Concat-LR")
            self.assertEqual(
                load_workbook(self._fp(td)).active.cell(2, 17).value, "Concat-LR")


# =============================================================================
# J  lr_onehot — compute_per_relation_metrics
# =============================================================================
class TestOneHot_PerRelationMetrics(unittest.TestCase):

    def test_perfect_macro_1(self):
        y = np.array(["A", "B", "C"])
        _, macro = _oh.compute_per_relation_metrics(y, y)
        self.assertAlmostEqual(macro, 1.0)

    def test_index_is_relation_name(self):
        y = np.array(["X", "Y"])
        df, _ = _oh.compute_per_relation_metrics(y, y)
        self.assertEqual(set(df.index), {"X", "Y"})

    def test_macro_is_mean_per_rel(self):
        y_t = np.array(["A", "A", "B", "B"])
        y_p = np.array(["A", "B", "B", "A"])
        df, macro = _oh.compute_per_relation_metrics(y_t, y_p)
        self.assertAlmostEqual(macro, df["hits_at_1"].mean(), places=3)

    def test_unseen_relation_hits_zero(self):
        y_t = np.array(["A", "A"])
        y_p = np.array(["B", "B"])
        df, _ = _oh.compute_per_relation_metrics(
            y_t, y_p, relations=["A", "B", "C"])
        self.assertEqual(df.loc["C", "hits_at_1"], 0.0)

    def test_rounded_to_4_places(self):
        y_t = np.array(["A"] * 7)
        y_p = np.array(["A"] * 3 + ["B"] * 4)
        df, _ = _oh.compute_per_relation_metrics(y_t, y_p)
        hits = df.loc["A", "hits_at_1"]
        self.assertEqual(hits, round(hits, 4))

    def test_count_column_correct(self):
        y_t = np.array(["A", "A", "B"])
        y_p = np.array(["A", "A", "B"])
        df, _ = _oh.compute_per_relation_metrics(y_t, y_p)
        self.assertEqual(df.loc["A", "count"], 2)

    def test_all_wrong_hits_zero(self):
        y_t = np.array(["A", "A"])
        y_p = np.array(["B", "B"])
        df, _ = _oh.compute_per_relation_metrics(y_t, y_p)
        self.assertEqual(df.loc["A", "hits_at_1"], 0.0)


# =============================================================================
# K  lr_onehot — record_results
# =============================================================================
class TestOneHot_RecordResults(unittest.TestCase):

    def _fp(self, td): return str(Path(td) / "oh.xlsx")

    def test_creates_file(self):
        with tempfile.TemporaryDirectory() as td:
            _oh.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            self.assertTrue(Path(self._fp(td)).exists())

    def test_default_eval_label_onehot_lr(self):
        with tempfile.TemporaryDirectory() as td:
            _oh.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            ws = load_workbook(self._fp(td)).active
            self.assertEqual(ws.cell(2, 17).value, "OneHot-LR")

    def test_appends_second_call(self):
        with tempfile.TemporaryDirectory() as td:
            _oh.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            _oh.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            self.assertEqual(load_workbook(self._fp(td)).active.max_row, 3)

    def test_mrr_col_18(self):
        with tempfile.TemporaryDirectory() as td:
            _oh.record_results(
                self._fp(td), _make_settings(), _make_metrics(mrr=0.55), 0)
            self.assertAlmostEqual(
                load_workbook(self._fp(td)).active.cell(2, 18).value, 0.55)

    def test_elapsed_formatted(self):
        with tempfile.TemporaryDirectory() as td:
            _oh.record_results(
                self._fp(td), _make_settings(), _make_metrics(), 7322)
            self.assertEqual(
                load_workbook(self._fp(td)).active.cell(2, 16).value, "2h 2m 2s")

    def test_freeze_pane_A2(self):
        with tempfile.TemporaryDirectory() as td:
            _oh.record_results(self._fp(td), _make_settings(), _make_metrics(), 0)
            self.assertEqual(
                str(load_workbook(self._fp(td)).active.freeze_panes), "A2")

    def test_custom_eval_label(self):
        with tempfile.TemporaryDirectory() as td:
            _oh.record_results(
                self._fp(td), _make_settings(), _make_metrics(), 0, "OHE-Test")
            self.assertEqual(
                load_workbook(self._fp(td)).active.cell(2, 17).value, "OHE-Test")


# =============================================================================
# L  caliensemble — fit_calibrators
# =============================================================================
class TestFitCalibrators(unittest.TestCase):

    def _balanced(self, N=90, R=3):
        """Balanced scores and labels for R relations."""
        rel_ids = np.array([i % R for i in range(N)])
        scores  = np.random.default_rng(0).standard_normal((N, R)).astype(np.float32)
        return scores, rel_ids

    def test_returns_list_of_len_num_relations(self):
        scores, rel_ids = self._balanced()
        cals = _cali.fit_calibrators(scores, rel_ids, 3)
        self.assertEqual(len(cals), 3)

    def test_degenerate_class_returns_none(self):
        # All labels are class 0 → binary labels for class 1 are all-zero → degenerate
        N, R = 50, 2
        rel_ids = np.zeros(N, dtype=int)
        scores  = np.random.default_rng(1).standard_normal((N, R)).astype(np.float32)
        cals = _cali.fit_calibrators(scores, rel_ids, R)
        self.assertIsNone(cals[1])

    def test_balanced_calibrators_not_none(self):
        scores, rel_ids = self._balanced(N=90, R=3)
        cals = _cali.fit_calibrators(scores, rel_ids, 3)
        self.assertTrue(all(c is not None for c in cals))

    def test_fitted_calibrator_has_predict_proba(self):
        scores, rel_ids = self._balanced()
        cals = _cali.fit_calibrators(scores, rel_ids, 3)
        for cal in cals:
            if cal is not None:
                self.assertTrue(hasattr(cal, "predict_proba"))

    def test_length_matches_num_relations_varied(self):
        for R in (2, 4):
            N = R * 30
            rel_ids = np.array([i % R for i in range(N)])
            scores  = np.random.default_rng(R).standard_normal((N, R)).astype(np.float32)
            cals = _cali.fit_calibrators(scores, rel_ids, R)
            self.assertEqual(len(cals), R)

    def test_predict_proba_output_shape(self):
        scores, rel_ids = self._balanced()
        cals = _cali.fit_calibrators(scores, rel_ids, 3)
        for cal in cals:
            if cal is not None:
                out = cal.predict_proba(scores[:, :1])
                self.assertEqual(out.shape[1], 2)


# =============================================================================
# M  caliensemble — apply_calibrators
# =============================================================================
class TestApplyCalibrators(unittest.TestCase):

    def _fitted(self, R=3):
        N = R * 30
        rel_ids = np.array([i % R for i in range(N)])
        scores  = np.random.default_rng(42).standard_normal((N, R)).astype(np.float32)
        return _cali.fit_calibrators(scores, rel_ids, R), R

    def test_output_shape(self):
        cals, R = self._fitted()
        scores  = np.random.default_rng(0).standard_normal((20, R))
        out = _cali.apply_calibrators(cals, scores, R)
        self.assertEqual(out.shape, (20, R))

    def test_rows_sum_to_one(self):
        cals, R = self._fitted()
        scores  = np.random.default_rng(1).standard_normal((20, R))
        out = _cali.apply_calibrators(cals, scores, R)
        np.testing.assert_allclose(out.sum(axis=1), np.ones(20), atol=1e-6)

    def test_all_values_in_0_to_1(self):
        cals, R = self._fitted()
        scores  = np.random.default_rng(2).standard_normal((20, R))
        out = _cali.apply_calibrators(cals, scores, R)
        self.assertTrue((out >= 0).all())
        self.assertTrue((out <= 1).all())

    def test_none_calibrator_uses_sigmoid(self):
        # sigmoid(0) = 0.5 on both columns → rows sum to 1 after normalisation
        cals   = [None, None]
        scores = np.array([[0.0, 0.0]])
        out    = _cali.apply_calibrators(cals, scores, 2)
        self.assertEqual(out.shape, (1, 2))
        self.assertAlmostEqual(out.sum(), 1.0, places=5)

    def test_output_dtype_float64(self):
        cals, R = self._fitted()
        scores  = np.random.default_rng(3).standard_normal((10, R))
        out = _cali.apply_calibrators(cals, scores, R)
        self.assertEqual(out.dtype, np.float64)

    def test_zero_rows_handled(self):
        # Edge case: if all sigmoid values somehow become 0, denominator is clamped to 1
        cals   = [None, None]
        scores = np.array([[-1000.0, -1000.0]])  # sigmoid ≈ 0 for both
        out    = _cali.apply_calibrators(cals, scores, 2)
        self.assertFalse(np.isnan(out).any())
        self.assertFalse(np.isinf(out).any())


# =============================================================================
# N  caliensemble — compute_metrics
# =============================================================================
class TestComputeMetrics(unittest.TestCase):

    REL_MAP = {0: "CONTRARY_TO", 1: "NOT_CONTRARY", 2: "SUPPORT"}

    def _perfect(self, true_ids, R=3):
        N = len(true_ids)
        p = np.zeros((N, R))
        for i, r in enumerate(true_ids):
            p[i, r] = 1.0
        return p

    def test_returns_overall_and_per_relation(self):
        ids  = np.array([0, 1, 2, 0])
        out  = _cali.compute_metrics(self._perfect(ids), ids, self.REL_MAP)
        self.assertIn("overall", out)
        self.assertIn("per_relation", out)

    def test_perfect_mrr_is_1(self):
        ids = np.array([0, 1, 2])
        out = _cali.compute_metrics(self._perfect(ids), ids, self.REL_MAP)
        self.assertAlmostEqual(out["overall"]["MRR"], 1.0)

    def test_perfect_micro_h1_is_1(self):
        ids = np.array([0, 1, 2, 0])
        out = _cali.compute_metrics(self._perfect(ids), ids, self.REL_MAP)
        self.assertAlmostEqual(out["overall"]["MicroH@1"], 1.0)

    def test_mrr_in_0_to_1(self):
        rng  = np.random.default_rng(7)
        ids  = np.array([0, 1, 2, 0, 1, 2])
        probs = rng.dirichlet([1] * 3, size=6)
        out  = _cali.compute_metrics(probs, ids, self.REL_MAP)
        self.assertGreater(out["overall"]["MRR"], 0)
        self.assertLessEqual(out["overall"]["MRR"], 1.0)

    def test_hits_monotone_h1_lte_h3(self):
        rng   = np.random.default_rng(8)
        ids   = np.array([0, 1, 2] * 5)
        probs = rng.dirichlet([1] * 3, size=15)
        out   = _cali.compute_metrics(probs, ids, self.REL_MAP)
        self.assertLessEqual(out["overall"]["MicroH@1"], out["overall"]["Hits@3"])

    def test_hits_monotone_h3_lte_h10(self):
        rng   = np.random.default_rng(9)
        ids   = np.array([0, 1, 2] * 5)
        probs = rng.dirichlet([1] * 3, size=15)
        out   = _cali.compute_metrics(probs, ids, self.REL_MAP)
        self.assertLessEqual(out["overall"]["Hits@3"], out["overall"]["Hits@10"])

    def test_per_relation_keys(self):
        ids = np.array([0, 1, 2])
        out = _cali.compute_metrics(self._perfect(ids), ids, self.REL_MAP)
        for v in out["per_relation"].values():
            for k in ("MRR", "Hits@1", "Hits@3", "Hits@10", "MeanRank", "N"):
                self.assertIn(k, v)

    def test_macro_h1_is_mean_of_per_rel(self):
        ids = np.array([0, 0, 1, 1, 2, 2])
        out = _cali.compute_metrics(self._perfect(ids), ids, self.REL_MAP)
        per_h1 = [v["Hits@1"] for v in out["per_relation"].values()]
        self.assertAlmostEqual(out["overall"]["MacroH@1"], np.mean(per_h1), places=5)

    def test_n_count_correct(self):
        ids = np.array([0, 1, 2, 0])
        out = _cali.compute_metrics(self._perfect(ids), ids, self.REL_MAP)
        self.assertEqual(out["overall"]["N"], 4)

    def test_per_relation_n_correct(self):
        ids = np.array([0, 0, 1, 2])
        out = _cali.compute_metrics(self._perfect(ids), ids, self.REL_MAP)
        self.assertEqual(out["per_relation"]["CONTRARY_TO"]["N"], 2)
        self.assertEqual(out["per_relation"]["NOT_CONTRARY"]["N"], 1)
        self.assertEqual(out["per_relation"]["SUPPORT"]["N"], 1)

    def test_mean_rank_always_gte_1(self):
        rng   = np.random.default_rng(11)
        ids   = np.array([0, 1, 2] * 4)
        probs = rng.dirichlet([1] * 3, size=12)
        out   = _cali.compute_metrics(probs, ids, self.REL_MAP)
        self.assertGreaterEqual(out["overall"]["MeanRank"], 1.0)

    def test_absent_relation_not_in_per_rel(self):
        # Only classes 0 and 1 present — class 2 should be absent
        ids  = np.array([0, 0, 1, 1])
        out  = _cali.compute_metrics(self._perfect(ids), ids, self.REL_MAP)
        self.assertNotIn("SUPPORT", out["per_relation"])


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    unittest.main(verbosity=2)
