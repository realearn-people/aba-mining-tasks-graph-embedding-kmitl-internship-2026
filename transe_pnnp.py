import pandas as pd
import numpy as np
import torch
import time
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pykeen.pipeline import pipeline
from pykeen.triples import TriplesFactory
from pykeen.sampling import BasicNegativeSampler
from sklearn.decomposition import PCA

# ── File paths ────────────────────────────────────────────────────────────────
CSV_PATH     = "hotel_contrary_dataset_PNNP.csv"   # input dataset
OUTPUT_DIR   = Path("transe_pnnp_output")               # folder for saved outputs
RESULTS_FILE = "experiment_results.xlsx"           # auto-recorded experiment log
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Hyperparameters ───────────────────────────────────────────────────────────
DATASET       = "PNNP"      # dataset label for experiment record
EMBEDDING_DIM = 100        # size of each entity/relation vector
NUM_EPOCHS    = 1000       # max epochs — early stopper will stop earlier
BATCH_SIZE    = 256        # number of triples processed per update step
LEARNING_RATE = 0.0001     # how much embeddings shift per step (Adam optimizer)
MARGIN        = 1.0        # minimum score gap between positive and negative triples
TRAIN_RATIO   = 0.8        # 80% of triples used for training
VAL_RATIO     = 0.1        # 10% used for validation (early stopping monitor), remaining 10% used for final test evaluation
RANDOM_SEED   = 42         # fixes randomness for reproducibility
NUM_NEG       = 2          # negatives per positive triple
MODEL         = "TransE"   # rotation-based KGE model — handles symmetric relations
LOSS          = "NSSALoss" # Non-Saturating Self-Adversarial Loss weights harder negatives more — better for relation corruption
ADV_TEMP      = 0.5        # adversarial temperature for NSSALoss
CORRUPT       = "relation" # corruption scheme — only replace the relation

# ── Extended negative sampler ─────────────────────────────────────────────────
class ExtendedBasicNegativeSampler(BasicNegativeSampler):
    """
    Custom negative sampler that extends BasicNegativeSampler.
    Overrides corrupt_batch() to capture and save every negative triple
    generated during training to a CSV file.

    Each batch writes immediately to disk — no accumulation in RAM —
    to avoid memory overflow on large datasets.

    Based on PyKEEN GitHub issue #1124:
    https://github.com/pykeen/pykeen/issues/1124#issuecomment-1262011487
    """

    def __init__(self, *args, save_path=str(OUTPUT_DIR / "negative_samples.csv"), **kwargs):
        # Force relation-only corruption inside the class
        # Cannot rely on PyKEEN passing corruption_scheme correctly to custom classes
        kwargs['corruption_scheme'] = ['relation']
        super().__init__(*args, **kwargs)

        # Use object.__setattr__ to bypass PyTorch nn.Module's attribute interception
        # nn.Module overrides __setattr__ for parameter management
        # which would reject plain Python attributes like these
        object.__setattr__(self, 'save_path', save_path)
        object.__setattr__(self, 'header_written', False)  # tracks if CSV header exists
        object.__setattr__(self, 'batch_count', 0)         # counts batches processed

    def corrupt_batch(self, positive_batch: torch.LongTensor) -> torch.LongTensor:
        """
        Called automatically by PyKEEN for every batch during training.
        Generates negatives using parent class logic, then saves them to CSV.
        Returns negatives unchanged so training proceeds normally.
        """
        # Generate negatives using BasicNegativeSampler's random relation corruption
        negatives = super().corrupt_batch(positive_batch=positive_batch)

        # Increment batch counter using object.__setattr__ (same reason as __init__)
        object.__setattr__(self, 'batch_count', self.batch_count + 1)

        # Debug messages — confirm corrupt_batch is being called
        if self.batch_count == 1:
            print(f"[DEBUG] corrupt_batch called for the first time!")
            print(f"[DEBUG] Negative shape: {negatives.shape}")
        if self.batch_count % 100 == 0:
            print(f"[DEBUG] Batches captured so far: {self.batch_count}")

        try:
            # Convert integer IDs back to readable phrase labels
            neg_flat = negatives.detach().cpu().reshape(-1, 3)
            rows = []
            for t in neg_flat:
                try:
                    rows.append({
                        "head":     id_to_entity.get(t[0].item(), "unknown"),
                        "relation": id_to_relation.get(t[1].item(), "unknown"),
                        "tail":     id_to_entity.get(t[2].item(), "unknown"),
                        "label":    0,   # 0 = negative (fake) triple
                    })
                except Exception as row_err:
                    print(f"[WARN] Skipping one row in batch {self.batch_count}: {row_err}")
                    continue
            if rows:
                df_batch = pd.DataFrame(rows)
                df_batch.to_csv(
                    self.save_path,
                    mode="w" if not self.header_written else "a",
                    header=not self.header_written,
                    index=False,      
                )
                object.__setattr__(self, 'header_written', True)
                del df_batch
            del neg_flat, rows
        except Exception as batch_err:
            print(f"[WARN] Skipping batch {self.batch_count} due to error: {batch_err}")
        return negatives

# ══════════════════════════════════════════════════════════════════════════════
# Relation-rank evaluator
# ══════════════════════════════════════════════════════════════════════════════
class RelationRankEvaluator:
    """
    Custom evaluator aligned with relation-corruption training.
 
    For each test triple (h, r, t), scores ALL possible relations and
    ranks the true relation r among them. Directly answers:
    'Given these two phrases, which relation holds between them?'
    — which is the core contrary mining question.
 
    Return format mirrors PyKEEN's nested metric dict structure so it
    can be used interchangeably with the default RankBasedEvaluator output.
 
    Outputs:
        metrics_dict : nested dict matching PyKEEN format
                       metrics_dict["both"]["realistic"]["inverse_harmonic_mean_rank"]
        df_eval      : per-triple DataFrame with relation_rank and hit columns
    """
 
    def __init__(self, model, triples_factory, batch_size=256, device=None):
        self.model         = model
        self.num_relations = triples_factory.num_relations
        self.batch_size    = batch_size
        self.device        = device or ("cuda" if torch.cuda.is_available() else "cpu")
 
    def evaluate(self, mapped_triples, additional_filter_triples=None):
        """
        Evaluate model using relation ranking.
 
        Args:
            mapped_triples            : test triples as integer tensor (N, 3)
            additional_filter_triples : list of tensors for filtered evaluation
                                        removes known true relations from ranking
 
        Returns:
            metrics_dict : nested dict — same format as PyKEEN's .to_dict()
            df_eval      : DataFrame — one row per test triple with rank details
        """
        self.model.eval()
        self.model.to(self.device)
 
        # Build filter map: (h_id, t_id) → set of all known true relation IDs
        # Used to exclude other known true relations when ranking (filtered eval)
        filter_map = {}
        if additional_filter_triples:
            for triples in additional_filter_triples:
                for triple in triples:
                    h, r, t = triple[0].item(), triple[1].item(), triple[2].item()
                    filter_map.setdefault((h, t), set()).add(r)
 
        eval_rows = []
        print(f"  Evaluating {len(mapped_triples):,} test triples...")
 
        with torch.no_grad():
            for i in range(0, len(mapped_triples), self.batch_size):
                batch = mapped_triples[i : i + self.batch_size]
 
                for triple in batch:
                    h_id = triple[0].item()
                    r_id = triple[1].item()
                    t_id = triple[2].item()
 
                    # Score all possible relations for this (h, ?, t) query
                    all_rels = torch.arange(self.num_relations, device=self.device)
                    h_tensor = torch.full((self.num_relations,), h_id, device=self.device)
                    t_tensor = torch.full((self.num_relations,), t_id, device=self.device)
                    hrt_all  = torch.stack([h_tensor, all_rels, t_tensor], dim=1)
                    scores   = self.model.score_hrt(hrt_all).squeeze()
 
                    # Filtered evaluation — mask out other known true relations
                    # for this (h, t) pair so they don't unfairly lower the rank
                    masked_scores = scores.clone()
                    known_rels    = filter_map.get((h_id, t_id), set())
                    for r_other in known_rels:
                        if r_other != r_id:
                            masked_scores[r_other] = float('-inf')
 
                    # Rank = number of relations scoring strictly higher + 1
                    true_score = masked_scores[r_id].item()
                    rank       = int((masked_scores > true_score).sum().item()) + 1
 
                    eval_rows.append({
                        "head":            id_to_entity.get(h_id, "unknown"),
                        "relation":        id_to_relation.get(r_id, "unknown"),
                        "tail":            id_to_entity.get(t_id, "unknown"),
                        "relation_rank":   rank,
                        "reciprocal_rank": round(1.0 / rank, 6),
                        "hits_at_1":       int(rank <= 1),
                        "hits_at_3":       int(rank <= 3),
                        "hits_at_10":      int(rank <= 10),
                    })
 
                # Progress update every 20 batches
                if (i // self.batch_size + 1) % 20 == 0:
                    done = min(i + self.batch_size, len(mapped_triples))
                    print(f"  Processed {done:,} / {len(mapped_triples):,}")
 
        # Compute aggregate metrics
        n      = len(eval_rows)
        mrr    = sum(r["reciprocal_rank"] for r in eval_rows) / n
        h_at_1 = sum(r["hits_at_1"]       for r in eval_rows) / n
        h_at_3 = sum(r["hits_at_3"]       for r in eval_rows) / n
        h_at_10= sum(r["hits_at_10"]      for r in eval_rows) / n
        mr     = sum(r["relation_rank"]   for r in eval_rows) / n
 
        # Wrap in PyKEEN-compatible nested dict format
        # mirrors: result.metric_results.to_dict()["both"]["realistic"]
        metrics_dict = {
            "both": {
                "realistic": {
                    "inverse_harmonic_mean_rank": round(mrr,    4),  # MRR
                    "arithmetic_mean_rank":       round(mr,     4),  # Mean Rank
                    "hits_at_1":                  round(h_at_1, 4),
                    "hits_at_3":                  round(h_at_3, 4),
                    "hits_at_10":                 round(h_at_10,4),
                }
            }
        }
 
        return metrics_dict, pd.DataFrame(eval_rows)
    
# ── Experiment recorder ───────────────────────────────────────────────────────
def record_results(filepath, settings, metrics, elapsed, eval_label="PyKEEN"):
    """
    Appends one row of settings and metrics to the experiment Excel file.
    Creates the file with a formatted header if it doesn't exist yet.
    Run number is auto-incremented based on existing rows.
    """
    headers = [
        "No.", "Timestamp", "Dataset", "Model", "Loss Function",
        "Neg Sampler", "Corruption", "Embedding Dim", "Num Epochs",
        "Best Epoch", "Batch Size", "Learning Rate", "Margin",
        "Adv Temp", "Num Neg", "Training Time", "Evaluator",
        "MRR", "Hits@1", "Hits@3", "Hits@10", "Mean Rank",
    ]

    header_fill  = PatternFill("solid", fgColor="0D1B4B")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Load existing file or create new one
    if Path(filepath).exists():
        wb       = load_workbook(filepath)
        ws       = wb.active
        next_row = ws.max_row + 1
        run_num  = next_row - 1   # row 1 is the header
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Experiment Results"
        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"    # keep header visible when scrolling
        next_row = 2
        run_num  = 1

    # Build the row with all settings and metrics
    row_data = [
        run_num,
        datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        settings["dataset"],
        settings["model"],
        settings["loss_func"],
        settings["neg_sampler"],
        settings["corrupt"],
        settings["embedding_dim"],
        settings["num_epochs"],
        settings.get("best_epoch", "—"),
        settings["batch_size"],
        settings["learning_rate"],
        settings["margin"],
        settings["adv_temp"],
        settings["num_neg"],
        f"{int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s",
        eval_label,
        round(metrics["mrr"],        4),
        round(metrics["hits_at_1"],  4),
        round(metrics["hits_at_3"],  4),
        round(metrics["hits_at_10"], 4),
        round(metrics["mean_rank"],  4),
    ]

    for col, val in enumerate(row_data, 1):
        ws.cell(row=next_row, column=col, value=val)

    wb.save(filepath)
    print(f"Recorded: {filepath}  Run #{run_num}")


# ══════════════════════════════════════════════════════════════════════════════
# Load data
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Loading triples")
print("=" * 60)

df = pd.read_csv(CSV_PATH)
print(f"Total rows loaded: {len(df):,}")
print(f"Relations found: {df['relation'].value_counts().to_dict()}")

# Extract only the 3 columns PyKEEN needs — head, relation, tail
triples_array = df[['head', 'relation', 'tail']].values.astype(str)

# ══════════════════════════════════════════════════════════════════════════════
# Build TriplesFactory and split
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Creating TriplesFactory")
print("=" * 60)

# TriplesFactory converts string phrases to integer IDs for the model
tf = TriplesFactory.from_labeled_triples(triples=triples_array)
print(f"Unique entities  : {tf.num_entities:,}")
print(f"Unique relations : {tf.num_relations}")

# Split into train / validation / test sets
# Test set = remaining 10% after train (80%) and val (10%)
training, validation, testing = tf.split(
    ratios=[TRAIN_RATIO, VAL_RATIO],
    random_state=RANDOM_SEED,
)
print(f"Train : {training.num_triples:,}")
print(f"Val   : {validation.num_triples:,}")
print(f"Test  : {testing.num_triples:,}")

# Build lookup dictionaries — used inside ExtendedBasicNegativeSampler
# to convert integer IDs back to readable phrase labels when saving to CSV
id_to_entity   = {v: k for k, v in training.entity_to_id.items()}
id_to_relation = {v: k for k, v in training.relation_to_id.items()}

# ══════════════════════════════════════════════════════════════════════════════
# Train
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Training " + MODEL)
print("=" * 60)

start_time = time.time()

result = pipeline(
    training=training,
    validation=validation,
    testing=testing,

    # Model — RotatE represents relations as rotations in complex space
    # Naturally handles symmetric relations like CONTRARY_TO
    model=MODEL,
    model_kwargs=dict(embedding_dim=EMBEDDING_DIM),

    # Loss — NSSALoss weights harder negatives more during training
    loss=LOSS,
    loss_kwargs=dict(
        margin=MARGIN,
        adversarial_temperature=ADV_TEMP,
    ),

    # Negative sampler — custom class that saves negatives to CSV during training
    # Pass class (not instance) — PyKEEN instantiates it internally
    # corruption_scheme is forced to ["relation"] inside the class __init__
    negative_sampler=ExtendedBasicNegativeSampler,
    negative_sampler_kwargs=dict(
        num_negs_per_pos=NUM_NEG,
        corruption_scheme=[CORRUPT],   # also passed here as documentation
    ),

    # Optimizer — Adam adapts learning rate per parameter
    # handles sparse gradients well (most entities appear rarely)
    optimizer="Adam",
    optimizer_kwargs=dict(lr=LEARNING_RATE),

    # Training loop — sLCWA samples NUM_NEG negatives per positive per batch
    # more efficient than full LCWA which scores all possible entities
    training_loop="sLCWA",
    training_kwargs=dict(
        num_epochs=NUM_EPOCHS,
        batch_size=BATCH_SIZE,
    ),

    # Early stopping — monitors Hits@10 on validation set every 10 epochs
    # stops if no improvement of at least 0.2% for 10 consecutive checks (100 epochs)
    stopper="early",
    stopper_kwargs=dict(
        metric="hits_at_10",
        patience=10,
        relative_delta=0.002,
        frequency=10,
    ),

    # Evaluation — rank-based: for each test triple, ranks all entities
    # filtered=True removes other known true triples from ranking for fairness
    evaluator="RankBasedEvaluator",
    evaluator_kwargs=dict(filtered=True),

    random_seed=RANDOM_SEED,

    # Automatically use GPU if available, otherwise fall back to CPU
    device="cuda" if torch.cuda.is_available() else "cpu",
)

end_time = time.time()
elapsed  = end_time - start_time
print(f"\nTraining completed in: {int(elapsed//3600)}h {int((elapsed%3600)//60)}m {int(elapsed%60)}s")

# ══════════════════════════════════════════════════════════════════════════════
# PyKEEN default evaluation results 
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Evaluation Results (PyKEEN default)")
print("=" * 60)
 
metrics_dict = result.metric_results.to_dict()
r_default    = metrics_dict["both"]["realistic"]
 
mrr        = r_default['inverse_harmonic_mean_rank']
hits_at_1  = r_default['hits_at_1']
hits_at_3  = r_default['hits_at_3']
hits_at_10 = r_default['hits_at_10']
mean_rank  = r_default['arithmetic_mean_rank']
 
print(f"MRR       : {mrr:.4f}")
print(f"Hits@1    : {hits_at_1:.4f}")
print(f"Hits@3    : {hits_at_3:.4f}")
print(f"Hits@10   : {hits_at_10:.4f}")
print(f"Mean Rank : {mean_rank:.4f}")

# ══════════════════════════════════════════════════════════════════════════════
# Record results to Excel
# ══════════════════════════════════════════════════════════════════════════════

# Get the epoch where the best validation result occurred
try:
    best_epoch = result.stopper.best_epoch
except Exception:
    best_epoch = "—"

settings = {
    "dataset":       DATASET,
    "model":         MODEL,
    "loss_func":     LOSS,
    "neg_sampler":   "ExtendedBasicNegativeSampler",
    "corrupt":       CORRUPT,
    "embedding_dim": EMBEDDING_DIM,
    "num_epochs":    NUM_EPOCHS,
    "best_epoch":    best_epoch,
    "batch_size":    BATCH_SIZE,
    "learning_rate": LEARNING_RATE,
    "margin":        MARGIN,
    "adv_temp":      ADV_TEMP,
    "num_neg":       NUM_NEG,
}

metrics_out = {
    "mrr":        mrr,
    "hits_at_1":  hits_at_1,
    "hits_at_3":  hits_at_3,
    "hits_at_10": hits_at_10,
    "mean_rank":  mean_rank,
}

record_results(RESULTS_FILE, settings, metrics_out, elapsed, eval_label="PyKEEN")

# ══════════════════════════════════════════════════════════════════════════════
# Relation-rank evaluation
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Relation-Rank Evaluation (aligned with relation corruption)")
print("=" * 60)
 
try:
    rel_evaluator = RelationRankEvaluator(
        model=result.model,
        triples_factory=training,
        batch_size=BATCH_SIZE,
        device="cuda" if torch.cuda.is_available() else "cpu",
    )
 
    # evaluate() returns same nested dict format as PyKEEN's .to_dict()
    rel_metrics_dict, df_rel_eval = rel_evaluator.evaluate(
        mapped_triples=testing.mapped_triples,
        additional_filter_triples=[
            training.mapped_triples,
            validation.mapped_triples,
        ],
    )
 
    # Access using same pattern as PyKEEN default evaluation above
    r_rel = rel_metrics_dict["both"]["realistic"]
 
    rel_mrr    = r_rel["inverse_harmonic_mean_rank"]
    rel_h1     = r_rel["hits_at_1"]
    rel_h3     = r_rel["hits_at_3"]
    rel_h10    = r_rel["hits_at_10"]
    rel_mr     = r_rel["arithmetic_mean_rank"]
 
    # Print side-by-side comparison — relation evaluator vs PyKEEN default
    print(f"\n{'Metric':<12} {'RelationRank':>14}  {'PyKEEN (entity)':>16}")
    print("-" * 46)
    print(f"{'MRR':<12} {rel_mrr:>14.4f}  {mrr:>16.4f}")
    print(f"{'Hits@1':<12} {rel_h1:>14.4f}  {hits_at_1:>16.4f}")
    print(f"{'Hits@3':<12} {rel_h3:>14.4f}  {hits_at_3:>16.4f}")
    print(f"{'Hits@10':<12} {rel_h10:>14.4f}  {hits_at_10:>16.4f}")
    print(f"{'Mean Rank':<12} {rel_mr:>14.4f}  {mean_rank:>16.4f}")
 
    # Record relation-rank results to Excel as a separate row
    metrics_rel_out = {
        "mrr":        rel_mrr,
        "hits_at_1":  rel_h1,
        "hits_at_3":  rel_h3,
        "hits_at_10": rel_h10,
        "mean_rank":  rel_mr,
    }
    record_results(
        RESULTS_FILE, settings, metrics_rel_out, elapsed,
        eval_label="RelationRank"
    )
 
    # Save per-triple relation ranks — overwrites each run
    rel_eval_path = OUTPUT_DIR / "relation_eval_ranked.csv"
    df_rel_eval.to_csv(rel_eval_path, index=False, mode='w')
    print(f"\nSaved: {rel_eval_path}  ({len(df_rel_eval):,} triples)")
 
except Exception as e:
    print(f"[ERROR] Relation-rank evaluation failed: {e}")
    import traceback; traceback.print_exc() 
 
 
# ══════════════════════════════════════════════════════════════════════════════
# Export visualization CSV
# RotatE embeddings are complex — concatenate real + imaginary for PCA
# ══════════════════════════════════════════════════════════════════════════════
print("=" * 60)
print("Exporting visualization data")
print("=" * 60)
 
try:
    # TransE embeddings are already real-valued — no complex conversion needed
    entity_emb_real = result.model.entity_representations[0](indices=None)\
                      .detach().cpu().numpy().astype(np.float32)
 
    print(f"Entity embedding shape (real+imag): {entity_emb_real.shape}")
 
    # Project to 2D with PCA
    pca       = PCA(n_components=2, random_state=RANDOM_SEED)
    coords_2d = pca.fit_transform(entity_emb_real)
    print(f"PCA explained variance ratio: {pca.explained_variance_ratio_.round(4)}")
 
    # Build entity dataframe with 2D coordinates
    entity_to_id = training.entity_to_id
    entity_rows  = []
    for entity_name, entity_id in entity_to_id.items():
        entity_rows.append({
            "entity_id":   entity_id,
            "entity_name": entity_name,
            "x":           round(float(coords_2d[entity_id, 0]), 6),
            "y":           round(float(coords_2d[entity_id, 1]), 6),
        })
    df_entities = pd.DataFrame(entity_rows)
 
    # ── Score all original triples ────────────────────────────────────────────
    df_orig = pd.read_csv(CSV_PATH)[['head', 'relation', 'tail']]
 
    scores_list = []
    result.model.eval()
    with torch.no_grad():
        for _, row in df_orig.iterrows():
            try:
                h_id = training.entity_to_id.get(str(row['head']), -1)
                r_id = training.relation_to_id.get(str(row['relation']), -1)
                t_id = training.entity_to_id.get(str(row['tail']), -1)
                if -1 in (h_id, r_id, t_id):
                    scores_list.append(None)
                    continue
                hrt = torch.tensor([[h_id, r_id, t_id]], dtype=torch.long)
                s   = result.model.score_hrt(hrt)
                scores_list.append(round(float(s.item()), 6))
            except Exception:
                scores_list.append(None)
 
    df_orig['score'] = scores_list
 
    # ── Join 2D coordinates for head and tail ─────────────────────────────────
    coord_map         = df_entities.set_index('entity_name')[['x', 'y']]
    df_orig['head_x'] = df_orig['head'].map(coord_map['x'])
    df_orig['head_y'] = df_orig['head'].map(coord_map['y'])
    df_orig['tail_x'] = df_orig['tail'].map(coord_map['x'])
    df_orig['tail_y'] = df_orig['tail'].map(coord_map['y'])
 
    # ── Save both visualization files ─────────────────────────────────────────
    VIZ_FILE    = OUTPUT_DIR / "visualization_data.csv"
    ENTITY_FILE = OUTPUT_DIR / "entity_embeddings.csv"
 
    df_orig.to_csv(VIZ_FILE, index=False)
    df_entities.to_csv(ENTITY_FILE, index=False)
 
    print(f"Saved: {VIZ_FILE}     ({len(df_orig):,} triples)")
    print(f"Saved: {ENTITY_FILE}  ({len(df_entities):,} entities)")
 
except Exception as e:
    print(f"[ERROR] Visualization export failed: {e}")
    import traceback; traceback.print_exc()
 
 
print("=" * 60)
print("All done.")
print("=" * 60)